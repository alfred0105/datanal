"""
화학실험 샘플 데이터 생성기
반응 합성 실험 데이터 (변수 12개, 80개 케이스)
"""
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

rng = np.random.RandomState(2024)

# ─── 실험 조건 정의 ──────────────────────────────────────────
N = 80  # 실험 케이스 수

# 각 변수: (이름, 단위, 최솟값, 최댓값, 최적구간 중심, 허용폭)
VARIABLES = [
    ("반응온도",     "°C",  50,  200, 130, 25),   # 최적 130±25
    ("반응압력",     "bar",  1,   20,  8,   3),    # 최적 8±3
    ("pH",          "",    2,   12,  7.5, 1.5),   # 최적 7.5±1.5
    ("촉매농도",     "g/L",  0.5,  5,  2.5, 0.8),  # 최적 2.5±0.8
    ("반응시간",     "min", 10,  180, 90,  25),    # 최적 90±25
    ("용매비율",     "%",   30,   90, 65,  10),    # 최적 65±10
    ("교반속도",     "rpm", 100, 600, 350, 80),    # 최적 350±80
    ("기질농도",     "mM",  10,  200, 80,  20),    # 최적 80±20
    ("질소유량",     "L/h",  0,   50, 25,   8),    # 최적 25±8
    ("냉각수온도",    "°C",  5,   30, 15,   5),    # 최적 15±5
    ("산화제농도",   "mM",  1,   30, 12,   4),     # 최적 12±4
    ("수분함량",     "%",   0.1, 10,  2,   1.5),   # 최적 2±1.5
]

VAR_NAMES = [v[0] for v in VARIABLES]
UNITS     = [v[1] for v in VARIABLES]

# ─── 실험 데이터 생성 ────────────────────────────────────────
def gen_experiment(case_id, success_prob_target):
    """실험 케이스 하나 생성. 성공 확률에 따라 변수값을 최적점 근처/외부로 설정"""
    row = [f"EXP-{case_id:03d}"]
    score_total = 0

    for name, unit, lo, hi, opt, span in VARIABLES:
        if rng.random() < success_prob_target:
            # 최적 구간 근처
            val = rng.normal(opt, span * 0.6)
        else:
            # 랜덤 (최적 구간 외부 가능)
            val = rng.uniform(lo, hi)
        val = round(float(np.clip(val, lo, hi)), 2)
        row.append(val)

        # 최적 구간 내 점수 계산
        dist = abs(val - opt) / span
        score_total += max(0, 1 - dist)

    # 성공/실패 판정: 전체 점수 기반 + 노이즈
    avg_score = score_total / len(VARIABLES)
    threshold = 0.52 + rng.normal(0, 0.08)
    result = "성공" if avg_score > threshold else "실패"
    row.append(result)
    return row

rows = []
# 성공률이 높은 실험 40개 + 낮은 실험 40개
for i in range(1, 41):
    rows.append(gen_experiment(i, success_prob_target=0.78))
for i in range(41, 81):
    rows.append(gen_experiment(i, success_prob_target=0.32))

rng.shuffle(rows)  # 섞기

# ─── 엑셀 저장 ───────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "합성실험"

# 헤더
headers = ["실험ID"] + [f"{n}({u})" if u else n for n, u in zip(VAR_NAMES, UNITS)] + ["결과"]
ws.append(headers)

# 데이터
for row in rows:
    ws.append(row)

# 스타일
bd = Border(
    left=Side(border_style="thin", color="2D3561"),
    right=Side(border_style="thin", color="2D3561"),
    top=Side(border_style="thin", color="2D3561"),
    bottom=Side(border_style="thin", color="2D3561"),
)

for ri, row_cells in enumerate(ws.iter_rows(1, N+1, 1, len(headers))):
    for ci, cell in enumerate(row_cells):
        cell.border    = bd
        cell.alignment = Alignment(horizontal="center", vertical="center")

        if ri == 0:
            cell.font = Font(bold=True, color="A0C4FF", size=10)
            cell.fill = PatternFill("solid", fgColor="111830")
        else:
            result_val = ws.cell(ri+1, len(headers)).value
            if result_val == "성공":
                cell.fill = PatternFill("solid", fgColor="0D2318")
                if ci == len(headers)-1:
                    cell.font = Font(bold=True, color="34D399")
            else:
                cell.fill = PatternFill("solid", fgColor="23100D")
                if ci == len(headers)-1:
                    cell.font = Font(bold=True, color="F87171")

# 열 너비
ws.column_dimensions['A'].width = 11
for idx in range(2, len(headers)+1):
    ws.column_dimensions[get_column_letter(idx)].width = 13

# 행 높이
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18

# 시트 두 번째: 변수 설명
ws2 = wb.create_sheet("변수설명")
ws2.append(["변수명", "단위", "최솟값", "최댓값", "최적구간", "설명"])
desc = [
    ("반응온도",   "°C",  50,  200, "105~155°C", "합성 반응이 일어나는 반응기 내부 온도"),
    ("반응압력",   "bar",  1,   20, "5~11 bar",  "반응기 내 가압 조건"),
    ("pH",        "",    2,   12, "6.0~9.0",   "반응 용액의 산도/염기도"),
    ("촉매농도",   "g/L", 0.5,  5, "1.7~3.3",   "금속 촉매의 농도"),
    ("반응시간",   "min", 10,  180, "65~115 min","목표 전환율 도달까지 소요 시간"),
    ("용매비율",   "%",   30,   90, "55~75%",    "유기용매 대 수상의 비율"),
    ("교반속도",   "rpm", 100, 600, "270~430 rpm","반응기 교반기 회전수"),
    ("기질농도",   "mM",  10,  200, "60~100 mM", "출발 물질의 초기 농도"),
    ("질소유량",   "L/h",  0,   50, "17~33 L/h", "불활성 기체 퍼징 유량"),
    ("냉각수온도",  "°C",  5,   30, "10~20°C",   "반응기 재킷 냉각수 온도"),
    ("산화제농도", "mM",  1,   30, "8~16 mM",   "산화 반응에 사용되는 산화제 농도"),
    ("수분함량",   "%",  0.1,  10, "0.5~3.5%",  "반응계 내 수분 함량"),
]
for d in desc:
    ws2.append(list(d))

for ci, w in enumerate([14,8,10,10,14,40], 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

out = Path("화학실험_합성반응_데이터.xlsx")
wb.save(out)

# 통계 출력
df = pd.DataFrame(rows, columns=["ID"] + VAR_NAMES + ["결과"])
s  = (df["결과"] == "성공").sum()
f  = (df["결과"] == "실패").sum()
print(f"생성 완료: {out}")
print(f"  전체 {N}건 | 성공 {s}건 ({s/N*100:.1f}%) | 실패 {f}건 ({f/N*100:.1f}%)")
print(f"  변수 {len(VAR_NAMES)}개: {VAR_NAMES}")
