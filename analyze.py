"""
====================================================
  방사형 데이터 분석기 v2 - 성공/실패 기반 분석
  Radar Chart + Success/Fail Outcome Analyzer
====================================================
엑셀 형식:
  첫 행  = 헤더 (변수명들 + 결과 컬럼)
  각 행  = 케이스 (숫자 변수들 + 성공/실패)
  결과 컬럼 = "성공", "성공", "실패" 등 (자동 인식)

실행 방법: 분석실행.bat 더블클릭
결과물  : output/ 폴더 PNG + HTML 자동 저장
"""
import sys, io
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8','utf8'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import os, glob, math, webbrowser
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.font_manager as fm

# ── 한글 폰트 (맑은 고딕) ────────────────────────────────────
def _setup_font():
    for _p in [r"C:\Windows\Fonts\malgun.ttf",
                r"C:\Windows\Fonts\gulim.ttc"]:
        if os.path.exists(_p):
            fm.fontManager.addfont(_p)
            _n = fm.FontProperties(fname=_p).get_name()
            matplotlib.rcParams['font.family'] = _n
            matplotlib.rcParams['axes.unicode_minus'] = False
            return _n
    matplotlib.rcParams['axes.unicode_minus'] = False
    return 'DejaVu Sans'

_FONT = _setup_font()

# ─────────────────────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────────────────────
OUTPUT_DIR = Path("output")
COLORS = ["#4f8ef7","#a78bfa","#34d399","#fbbf24",
          "#f472b6","#38bdf8","#fb923c","#a3e635"]

# 성공을 나타내는 키워드 (대소문자 무시)
SUCCESS_KEYWORDS = {'성공','success','pass','합격','yes','y','true','1','ok','good'}
FAIL_KEYWORDS    = {'실패','fail','failure','불합격','no','n','false','0','bad','ng'}

BG   = "#0d0f1a"
CARD = "#12152b"
GRID = "#1e2240"
TEXT = "#c0c8f0"

# ─────────────────────────────────────────────────────────────
#  1. 샘플 엑셀 생성
# ─────────────────────────────────────────────────────────────
def create_sample_excel(path: Path):
    """성공/실패 컬럼이 포함된 샘플 데이터 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "데이터"

    headers = ["케이스","속도","정확도","안정성","효율성","비용절감","범위","결과"]
    rng = np.random.RandomState(42)

    # 성공 패턴: 속도·정확도·안정성이 높을수록 성공 확률 높음
    rows = []
    for i in range(30):
        vals = rng.randint(40, 100, 6).tolist()
        score = (vals[0]*0.25 + vals[1]*0.3 + vals[2]*0.2 +
                 vals[3]*0.1  + vals[4]*0.1  + vals[5]*0.05)
        result = "성공" if score + rng.normal(0,8) > 68 else "실패"
        rows.append([f"케이스{i+1:02d}"] + vals + [result])

    hfill = PatternFill("solid", fgColor="1E2240")
    hfont = Font(bold=True, color="7EB3FF", size=11)
    bd    = Border(left=Side(border_style="thin", color="2D3561"),
                   right=Side(border_style="thin", color="2D3561"),
                   top=Side(border_style="thin", color="2D3561"),
                   bottom=Side(border_style="thin", color="2D3561"))
    sfill = PatternFill("solid", fgColor="0D2B1A")
    ffill = PatternFill("solid", fgColor="2B0D0D")

    ws.append(headers)
    for row in rows:
        ws.append(row)

    for ri, row_cells in enumerate(ws.iter_rows(1, len(rows)+1, 1, len(headers))):
        for ci, cell in enumerate(row_cells):
            cell.border    = bd
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ri == 0:
                cell.font = hfont
                cell.fill = hfill
            else:
                result_val = ws.cell(ri+1, len(headers)).value
                if result_val == "성공":
                    cell.fill = sfill
                else:
                    cell.fill = ffill

    ws.column_dimensions['A'].width = 14
    for c in 'BCDEFGH':
        ws.column_dimensions[c].width = 10
    ws.column_dimensions['H'].width = 8

    wb.save(path)
    print(f"  [OK] 샘플 파일 생성: {path.name} ({len(rows)}개 케이스)")
    return path


# ─────────────────────────────────────────────────────────────
#  2. 엑셀 자동 로드 + 파싱
# ─────────────────────────────────────────────────────────────
def load_and_parse(filepath: str):
    """
    엑셀 자동 파싱:
    - 숫자 열 = 변수(축)
    - 성공/실패 문자열 열 = 결과 컬럼 (자동 감지)
    반환: df_vars (숫자 변수 DataFrame), labels (성공=True/실패=False), var_names
    """
    xl   = pd.ExcelFile(filepath)
    # 데이터가 가장 많은 시트 선택
    best, best_n = xl.sheet_names[0], -1
    for sn in xl.sheet_names:
        tmp = pd.read_excel(filepath, sheet_name=sn, header=None)
        n   = tmp.select_dtypes(include=[np.number]).size
        if n > best_n:
            best_n, best = n, sn
    print(f"  [OK] 시트 선택: '{best}'")

    df = pd.read_excel(filepath, sheet_name=best)

    # 결과 컬럼 자동 감지
    result_col = None
    for col in df.columns:
        uniq = df[col].dropna().astype(str).str.strip().str.lower().unique()
        matched = set(uniq) & (SUCCESS_KEYWORDS | FAIL_KEYWORDS)
        if len(matched) >= 1 and len(set(uniq) - SUCCESS_KEYWORDS - FAIL_KEYWORDS) == 0:
            result_col = col
            break

    if result_col is None:
        # 마지막 열이 문자열이면 결과 컬럼으로 사용
        last = df.columns[-1]
        if df[last].dtype == object:
            result_col = last

    if result_col is None:
        raise ValueError(
            "결과 컬럼(성공/실패)을 찾을 수 없습니다.\n"
            "엑셀에 '성공' 또는 '실패' 값이 있는 컬럼을 추가해주세요.")

    print(f"  [OK] 결과 컬럼 인식: '{result_col}'")

    # 레이블
    raw_labels = df[result_col].astype(str).str.strip().str.lower()
    labels = raw_labels.map(lambda v: True if v in SUCCESS_KEYWORDS else
                                      (False if v in FAIL_KEYWORDS else None))
    valid_mask = labels.notna()
    df     = df[valid_mask].reset_index(drop=True)
    labels = labels[valid_mask].reset_index(drop=True)

    success_n = int(labels.sum())
    fail_n    = int((~labels).sum())
    total     = len(labels)
    base_rate = round(success_n / total * 100, 1) if total > 0 else 0
    print(f"  [OK] 전체 {total}건: 성공 {success_n}건 / 실패 {fail_n}건  (기본율 {base_rate}%)")

    # 숫자 변수 컬럼만 추출
    num_cols = [c for c in df.columns
                if c != result_col
                and pd.api.types.is_numeric_dtype(df[c])]
    if not num_cols:
        # 숫자로 변환 가능한 열 찾기
        for c in df.columns:
            if c == result_col: continue
            try:
                df[c] = pd.to_numeric(df[c], errors='coerce')
                if df[c].notna().sum() > total * 0.5:
                    num_cols.append(c)
            except Exception:
                pass

    df_vars = df[num_cols].copy()
    print(f"  [OK] 변수(축) {len(num_cols)}개: {num_cols}")

    return df_vars, labels.astype(bool), num_cols, base_rate


# ─────────────────────────────────────────────────────────────
#  3. 축별 성공 확률 계산
# ─────────────────────────────────────────────────────────────
def calc_axis_probs(df_vars: pd.DataFrame, labels: pd.Series) -> dict:
    """
    각 변수(축)별로: 성공 케이스의 평균과 실패 케이스의 평균을 분리하고
    전체 값 범위에서 성공이 차지하는 비율을 성공 확률로 반환
    """
    results = {}
    for col in df_vars.columns:
        vals   = df_vars[col].values
        s_vals = vals[labels.values]
        f_vals = vals[~labels.values]

        s_mean = float(np.nanmean(s_vals)) if len(s_vals) > 0 else 0
        f_mean = float(np.nanmean(f_vals)) if len(f_vals) > 0 else 0
        total  = len(labels)
        s_count = int(labels.sum())

        # 성공 확률: 기본 성공률 × 이 변수에서 성공 그룹이 얼마나 높은 값인지
        base_rate = s_count / total if total > 0 else 0.5
        col_min   = float(np.nanmin(vals)) if len(vals) > 0 else 0
        col_max   = float(np.nanmax(vals)) if len(vals) > 0 else 1
        rng       = col_max - col_min if col_max > col_min else 1

        # 성공 그룹이 실패 그룹보다 높은 정도 (분리도)
        separation = (s_mean - f_mean) / rng  # -1 ~ +1

        # 성공 확률 = 기본 성공률 + 분리도 기여분
        prob = base_rate + separation * 0.3
        prob = round(max(5, min(95, prob * 100)))

        results[col] = {
            'prob'   : prob,
            's_mean' : s_mean,
            'f_mean' : f_mean,
            's_count': int(labels.sum()),
            'f_count': int((~labels).sum()),
            'col_min': col_min,
            'col_max': col_max,
        }
    return results


# ─────────────────────────────────────────────────────────────
#  4. 예측값 기반 성공 확률 계산 (KNN 거리 방식)
# ─────────────────────────────────────────────────────────────
def predict_success_prob(df_vars: pd.DataFrame, labels: pd.Series,
                         input_vals: list, k: int = 7) -> float:
    """
    예측값 리스트 input_vals (각 변수 순서)를 받아
    KNN (가중 거리) 방식으로 성공 확률 반환 (0~100%)
    """
    df_norm = df_vars.copy()
    input_norm = []
    for i, col in enumerate(df_vars.columns):
        mn, mx = df_norm[col].min(), df_norm[col].max()
        if mx > mn:
            df_norm[col] = (df_norm[col] - mn) / (mx - mn)
            input_norm.append((input_vals[i] - mn) / (mx - mn))
        else:
            df_norm[col] = 0.5
            input_norm.append(0.5)

    arr   = df_norm.values
    ivec  = np.array(input_norm)
    dists = np.sqrt(np.sum((arr - ivec) ** 2, axis=1)) + 1e-9

    # 가장 가까운 k개
    k     = min(k, len(dists))
    idx   = np.argsort(dists)[:k]
    w     = 1.0 / dists[idx]
    w    /= w.sum()
    prob  = float(np.sum(w * labels.values[idx].astype(float)))
    return round(prob * 100, 1)


# ─────────────────────────────────────────────────────────────
#  5. 방사형 차트 시각화
# ─────────────────────────────────────────────────────────────
def prob_color(prob: float):
    t = prob / 100.0
    if t < 0.5:
        r = 0.94; g = 0.30 + t * 0.78; b = 0.27
    else:
        t2 = (t - 0.5) * 2
        r  = 0.94 - t2 * 0.80; g = 0.69 + t2 * 0.07; b = 0.27 - t2 * 0.14
    return (max(0,min(1,r)), max(0,min(1,g)), max(0,min(1,b)))

def rgb_hex(rgb): return '#%02x%02x%02x' % tuple(int(c*255) for c in rgb)


def _radar_setup(ax, angles, ylim=100):
    """방사형 차트 공통 설정"""
    ax.set_facecolor(CARD)
    ax.set_theta_offset(np.pi / 2)
    ax.set_theta_direction(-1)
    ax.set_ylim(0, ylim)
    ax.set_yticks([20, 40, 60, 80, 100])
    ax.set_yticklabels(['20', '40', '60', '80', '100'],
                        color='#3a4270', fontsize=7, fontfamily=_FONT)
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels([])
    ax.grid(color=GRID, linewidth=0.7, linestyle='--', alpha=0.7)
    ax.spines['polar'].set_color(GRID)


def draw_radar(df_vars: pd.DataFrame, labels: pd.Series,
               axis_probs: dict, base_rate: float, output_path: Path):

    var_names = list(df_vars.columns)
    N         = len(var_names)
    angles    = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
    angles   += angles[:1]
    angles_arr = np.array(angles)

    # ── 정규화 (0~100%) ──────────────────────────────────────
    df_norm = df_vars.copy()
    for col in df_vars.columns:
        mn, mx = df_norm[col].min(), df_norm[col].max()
        df_norm[col] = (df_norm[col] - mn) / (mx - mn) * 100 if mx > mn else 50.0

    s_means = [float(df_norm.loc[labels,  c].mean()) for c in var_names]
    f_means = [float(df_norm.loc[~labels, c].mean()) for c in var_names]
    s_mins  = [float(df_norm.loc[labels,  c].min())  for c in var_names]
    s_maxs  = [float(df_norm.loc[labels,  c].max())  for c in var_names]
    f_mins  = [float(df_norm.loc[~labels, c].min())  for c in var_names]
    f_maxs  = [float(df_norm.loc[~labels, c].max())  for c in var_names]

    for lst in [s_means, f_means, s_mins, s_maxs, f_mins, f_maxs]:
        lst += lst[:1]

    s_n = int(labels.sum()); f_n = int((~labels).sum()); tot = len(labels)

    # ── Figure 구성 ──────────────────────────────────────────
    fig = plt.figure(figsize=(26, 14), facecolor=BG)

    # 타이틀
    fig.text(0.5, 0.97,
             "방사형 분석 · 성공 vs 실패 상세 비교",
             ha='center', fontsize=18, color='white',
             fontweight='bold', fontfamily=_FONT)
    fig.text(0.5, 0.945,
             f"전체 {tot}건  |  성공 {s_n}건  |  실패 {f_n}건  |  "
             f"기본 성공률 {base_rate:.1f}%  |  "
             f"{datetime.now().strftime('%Y-%m-%d %H:%M')}",
             ha='center', fontsize=11, color='#7080b8', fontfamily=_FONT)

    # ─────────────────────────────────────────────────────────
    # 패널 A: 평균 비교 (왼쪽)
    # ─────────────────────────────────────────────────────────
    axA = fig.add_axes([0.01, 0.08, 0.34, 0.84], polar=True)
    _radar_setup(axA, angles)

    # 성공 평균 - 채운 실선
    axA.fill(angles, s_means, color='#3b82f6', alpha=0.25, zorder=3)
    axA.plot(angles, s_means, color='#60a5fa', linewidth=2.8, zorder=5)
    axA.scatter(angles[:-1], s_means[:-1],
                color='#93c5fd', s=70, zorder=6,
                edgecolors='#1d4ed8', linewidths=1.5)

    # 실패 평균 - 채운 파선
    axA.fill(angles, f_means, color='#ef4444', alpha=0.18, zorder=3)
    axA.plot(angles, f_means, color='#f87171', linewidth=2.8,
             linestyle='--', zorder=5)
    axA.scatter(angles[:-1], f_means[:-1],
                color='#fca5a5', s=70, zorder=6,
                edgecolors='#991b1b', linewidths=1.5)

    # 축 라벨 + 성공 확률 뱃지
    label_r = 116
    for ai, vname in enumerate(var_names):
        prob = axis_probs[vname]['prob']
        rgb  = prob_color(prob)
        axA.text(angles[ai], label_r,
                 f"{vname}\n{prob}%",
                 ha='center', va='center', fontsize=7.5,
                 color=TEXT, fontweight='bold', fontfamily=_FONT,
                 bbox=dict(boxstyle='round,pad=0.28',
                           facecolor=rgb + (0.20,),
                           edgecolor=rgb + (0.7,), linewidth=0.8))

    # 패널 A 범례
    hA = [
        plt.Line2D([0],[0], color='#60a5fa', linewidth=2.8,
                   label=f'성공 평균  (n={s_n})'),
        plt.Line2D([0],[0], color='#f87171', linewidth=2.8,
                   linestyle='--', label=f'실패 평균  (n={f_n})'),
    ]
    axA.legend(handles=hA, loc='lower left',
               bbox_to_anchor=(-0.12, -0.08), framealpha=0.25,
               facecolor=CARD, edgecolor=GRID, labelcolor=TEXT,
               fontsize=9.5, prop={'family': _FONT})

    fig.text(0.175, 0.085, "① 평균값 비교",
             ha='center', fontsize=11, color='#7080b8',
             fontfamily=_FONT, fontstyle='italic')

    # ─────────────────────────────────────────────────────────
    # 패널 B: 분포 범위 (가운데)
    # ─────────────────────────────────────────────────────────
    axB = fig.add_axes([0.36, 0.08, 0.27, 0.84], polar=True)
    _radar_setup(axB, angles)

    # 성공 범위 밴드 (파랑, 실선 테두리)
    axB.fill_between(angles_arr, s_mins, s_maxs,
                     color='#3b82f6', alpha=0.30, zorder=2)
    axB.plot(angles, s_mins, color='#93c5fd', linewidth=1.4,
             linestyle='-', zorder=4)
    axB.plot(angles, s_maxs, color='#93c5fd', linewidth=1.4,
             linestyle='-', zorder=4)
    # 성공 평균선 (가이드용, 가는 흰선)
    axB.plot(angles, s_means, color='white', linewidth=1.2,
             linestyle=':', alpha=0.6, zorder=5)

    # 실패 범위 밴드 (주황, 빗금 패턴으로 겹침 구분)
    axB.fill_between(angles_arr, f_mins, f_maxs,
                     color='#f97316', alpha=0.22, zorder=3)
    axB.plot(angles, f_mins, color='#fdba74', linewidth=1.4,
             linestyle='-', zorder=4)
    axB.plot(angles, f_maxs, color='#fdba74', linewidth=1.4,
             linestyle='-', zorder=4)
    # 실패 평균선 (가이드용)
    axB.plot(angles, f_means, color='white', linewidth=1.2,
             linestyle=':', alpha=0.6, zorder=5)

    # 축 라벨
    for ai, vname in enumerate(var_names):
        axB.text(angles[ai], label_r, vname,
                 ha='center', va='center', fontsize=7.5,
                 color=TEXT, fontweight='bold', fontfamily=_FONT,
                 bbox=dict(boxstyle='round,pad=0.22',
                           facecolor='#12152b', edgecolor=GRID,
                           linewidth=0.6))

    # 패널 B 범례
    hB = [
        mpatches.Patch(color='#3b82f6', alpha=0.55,
                       label=f'성공 분포 범위  (n={s_n})'),
        plt.Line2D([0],[0], color='#93c5fd', linewidth=1.4,
                   label='성공 최솟값 / 최댓값'),
        mpatches.Patch(color='#f97316', alpha=0.45,
                       label=f'실패 분포 범위  (n={f_n})'),
        plt.Line2D([0],[0], color='#fdba74', linewidth=1.4,
                   label='실패 최솟값 / 최댓값'),
        plt.Line2D([0],[0], color='white', linewidth=1.2,
                   linestyle=':', alpha=0.7, label='각 그룹 평균 (점선)'),
    ]
    axB.legend(handles=hB, loc='lower left',
               bbox_to_anchor=(-0.18, -0.10), framealpha=0.25,
               facecolor=CARD, edgecolor=GRID, labelcolor=TEXT,
               fontsize=9.0, prop={'family': _FONT})

    fig.text(0.495, 0.085, "② 분포 범위 (파랑=성공 · 주황=실패)",
             ha='center', fontsize=11, color='#7080b8',
             fontfamily=_FONT, fontstyle='italic')

    # ─────────────────────────────────────────────────────────
    # 패널 C: 분석 테이블 (오른쪽)
    # ─────────────────────────────────────────────────────────
    ax_t = fig.add_axes([0.65, 0.08, 0.34, 0.84])
    ax_t.set_facecolor(CARD); ax_t.axis('off')

    col_hdrs = ["변수", "성공avg", "실패avg", "성공확률"]
    col_ws   = [2.0, 1.0, 1.0, 1.1]
    total_w  = sum(col_ws)
    col_xs   = [sum(col_ws[:i]) / total_w for i in range(len(col_ws))]
    n_rows   = len(var_names) + 2
    cell_h   = 0.88 / n_rows
    y_top    = 0.97

    def draw_cell(x, y, w, h, txt, bg, fg, bold=False, fs=12):
        ax_t.add_patch(mpatches.FancyBboxPatch(
            (x + .003, y + .003), w - .006, h - .006,
            boxstyle="round,pad=0.005", linewidth=0.5,
            edgecolor=GRID, facecolor=bg,
            transform=ax_t.transAxes, clip_on=False))
        ax_t.text(x + w / 2, y + h / 2, txt,
                  ha='center', va='center', fontsize=fs, color=fg,
                  fontweight='bold' if bold else 'normal',
                  transform=ax_t.transAxes, clip_on=False,
                  fontfamily=_FONT)

    # 헤더
    cws = [w / total_w for w in col_ws]
    for hdr, cx, cw in zip(col_hdrs, col_xs, cws):
        draw_cell(cx, y_top - cell_h, cw, cell_h,
                  hdr, "#1e2650", "#7eb3ff", bold=True, fs=11)

    # 데이터 행
    sorted_vars = sorted(var_names, key=lambda v: -axis_probs[v]['prob'])
    row_bgs = ["#12152b", "#161a30"]
    for ri, vname in enumerate(sorted_vars):
        info = axis_probs[vname]
        prob = info['prob']
        rgb  = prob_color(prob); hx = rgb_hex(rgb)
        y    = y_top - cell_h * (ri + 2)
        draw_cell(col_xs[0], y, cws[0], cell_h,
                  vname, row_bgs[ri % 2], TEXT, fs=11)
        draw_cell(col_xs[1], y, cws[1], cell_h,
                  f"{info['s_mean']:.1f}", row_bgs[ri % 2], '#60a5fa', fs=11)
        draw_cell(col_xs[2], y, cws[2], cell_h,
                  f"{info['f_mean']:.1f}", row_bgs[ri % 2], '#f87171', fs=11)
        draw_cell(col_xs[3], y, cws[3], cell_h,
                  f"{prob}%", hx + '22', hx, bold=True, fs=13)

    # 전체 요약
    y_ov   = y_top - cell_h * (len(var_names) + 2)
    ov_rgb = prob_color(base_rate); ov_hex = rgb_hex(ov_rgb)
    draw_cell(0, y_ov, 1.0, cell_h,
              f"기본 성공률 (전체): {base_rate:.1f}%",
              ov_hex + '22', ov_hex, bold=True, fs=14)

    ax_t.text(0.5, y_ov - 0.028,
              "성공확률 = 각 변수에서 성공 그룹이 실패 그룹보다\n"
              "높은 정도를 반영한 가중 추정값",
              ha='center', va='top', fontsize=10, color='#5c6490',
              transform=ax_t.transAxes, fontfamily=_FONT)

    fig.text(0.825, 0.085, "③ 변수별 성공 확률 (내림차순)",
             ha='center', fontsize=11, color='#7080b8',
             fontfamily=_FONT, fontstyle='italic')

    plt.savefig(output_path, dpi=155, bbox_inches='tight',
                facecolor=BG, edgecolor='none')
    plt.close(fig)
    print(f"  [OK] 차트 저장: {output_path}")

    var_names = list(df_vars.columns)
    N         = len(var_names)
    angles    = np.linspace(0, 2*np.pi, N, endpoint=False).tolist()
    angles   += angles[:1]

    # 정규화 평균값 (성공/실패 그룹별)
    df_norm = df_vars.copy()
    for col in df_vars.columns:
        mn, mx = df_norm[col].min(), df_norm[col].max()
        df_norm[col] = (df_norm[col]-mn)/(mx-mn)*100 if mx > mn else 50.0

    s_means = [float(df_norm.loc[labels,  c].mean()) for c in var_names]
    f_means = [float(df_norm.loc[~labels, c].mean()) for c in var_names]
    s_means += s_means[:1]; f_means += f_means[:1]

    fig = plt.figure(figsize=(22, 13), facecolor=BG)
    fig.subplots_adjust(left=0.01, right=0.99, top=0.91, bottom=0.05)

    # 타이틀
    fig.text(0.5, 0.96,
        "방사형 데이터 분석 · 성공/실패 분포 & 성공 확률",
        ha='center', fontsize=17, color='white', fontweight='bold',
        fontfamily=_FONT)
    s_n = int(labels.sum()); f_n = int((~labels).sum()); tot = len(labels)
    fig.text(0.5, 0.93,
        f"전체 {tot}건  |  성공 {s_n}건  |  실패 {f_n}건  |  "
        f"기본 성공률 {base_rate:.1f}%  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        ha='center', fontsize=10, color='#7080b8', fontfamily=_FONT)

    # ── 방사형 차트 ──────────────────────────────────────────
    ax = fig.add_axes([0.01, 0.06, 0.57, 0.85], polar=True)
    ax.set_facecolor(CARD)
    ax.set_theta_offset(np.pi/2); ax.set_theta_direction(-1)
    ax.set_ylim(0, 100)
    ax.set_yticks([20,40,60,80,100])
    ax.set_yticklabels(['20%','40%','60%','80%','100%'],
                        color='#4a5280', fontsize=8, fontfamily=_FONT)
    ax.set_xticks(angles[:-1]); ax.set_xticklabels([])
    ax.grid(color=GRID, linewidth=0.8, linestyle='--', alpha=0.8)
    ax.spines['polar'].set_color(GRID)

    # 정규화 min/max (성공/실패 그룹별)
    s_mins = [float(df_norm.loc[labels,  c].min()) for c in var_names]
    s_maxs = [float(df_norm.loc[labels,  c].max()) for c in var_names]
    f_mins = [float(df_norm.loc[~labels, c].min()) for c in var_names]
    f_maxs = [float(df_norm.loc[~labels, c].max()) for c in var_names]
    # 닫힌 다각형을 위해 첫값 반복
    s_mins += s_mins[:1]; s_maxs += s_maxs[:1]
    f_mins += f_mins[:1]; f_maxs += f_maxs[:1]

    angles_arr = np.array(angles)

    # ── 성공 그룹 (파란색 계열) ──────────────────────────────
    # ① 최솟값~최댓값 사이 범위 밴드 (아주 연한 파랑)
    ax.fill_between(angles_arr, s_mins, s_maxs,
                    color='#4f8ef7', alpha=0.12, zorder=2)
    # ② 최솟값 경계선 (가는 점선)
    ax.plot(angles, s_mins, color='#4f8ef7', linewidth=1.0,
            linestyle=':', alpha=0.6, zorder=4)
    # ③ 최댓값 경계선 (가는 점선)
    ax.plot(angles, s_maxs, color='#4f8ef7', linewidth=1.0,
            linestyle=':', alpha=0.6, zorder=4)
    # ④ 평균 실선 (굵고 불투명) → 가장 눈에 띄는 선
    ax.plot(angles, s_means, color='#2563eb', linewidth=3.0, zorder=6)
    # ⑤ 평균 점 (크고 밝은 원형)
    ax.scatter(angles[:-1], s_means[:-1],
               color='#2563eb', s=80, zorder=7, edgecolors='white', linewidths=1.5)
    # ⑥ 최솟값 ▼ 삼각형 점
    ax.scatter(angles[:-1], s_mins[:-1],
               color='#93c5fd', s=45, zorder=5, marker='v',
               edgecolors='#2563eb', linewidths=1.0)
    # ⑦ 최댓값 ▲ 삼각형 점
    ax.scatter(angles[:-1], s_maxs[:-1],
               color='#93c5fd', s=45, zorder=5, marker='^',
               edgecolors='#2563eb', linewidths=1.0)

    # ── 실패 그룹 (빨간색 계열) ──────────────────────────────
    # ① 최솟값~최댓값 범위 밴드 (아주 연한 빨강)
    ax.fill_between(angles_arr, f_mins, f_maxs,
                    color='#ef4444', alpha=0.09, zorder=2)
    # ② 최솟값 경계선
    ax.plot(angles, f_mins, color='#ef4444', linewidth=1.0,
            linestyle=':', alpha=0.6, zorder=4)
    # ③ 최댓값 경계선
    ax.plot(angles, f_maxs, color='#ef4444', linewidth=1.0,
            linestyle=':', alpha=0.6, zorder=4)
    # ④ 평균 점선 (굵은 파선)
    ax.plot(angles, f_means, color='#dc2626', linewidth=3.0,
            linestyle='--', zorder=6)
    # ⑤ 평균 점
    ax.scatter(angles[:-1], f_means[:-1],
               color='#dc2626', s=80, zorder=7, edgecolors='white', linewidths=1.5)
    # ⑥ 최솟값 ▼
    ax.scatter(angles[:-1], f_mins[:-1],
               color='#fca5a5', s=45, zorder=5, marker='v',
               edgecolors='#dc2626', linewidths=1.0)
    # ⑦ 최댓값 ▲
    ax.scatter(angles[:-1], f_maxs[:-1],
               color='#fca5a5', s=45, zorder=5, marker='^',
               edgecolors='#dc2626', linewidths=1.0)

    # 축 라벨 + 성공확률 뱃지
    for ai, vname in enumerate(var_names):
        info  = axis_probs[vname]
        prob  = info['prob']
        rgb   = prob_color(prob)
        hex_c = rgb_hex(rgb)
        label_r = 113

        ax.scatter([angles[ai]], [prob], color=[rgb], s=130,
                   zorder=8, edgecolors='white', linewidths=1.0, alpha=0.9)

        ax.text(angles[ai], label_r,
                f"{vname}\n{prob}%",
                ha='center', va='center', fontsize=8.5,
                color=TEXT, fontweight='bold', fontfamily=_FONT,
                bbox=dict(boxstyle='round,pad=0.3',
                          facecolor=rgb+(0.18,),
                          edgecolor=rgb+(0.6,), linewidth=0.7))

    # 범례 (명확한 설명 포함)
    legend_handles = [
        plt.Line2D([0],[0], color='#2563eb', linewidth=3.0,
                   label=f'● 성공 평균선 (n={s_n})'),
        mpatches.Patch(color='#4f8ef7', alpha=0.3,
                       label='  성공 범위 (최솟값~최댓값)'),
        plt.Line2D([0],[0], color='#4f8ef7', linewidth=1.0, linestyle=':',
                   label='  ···  성공 최솟값/최댓값 경계'),
        plt.Line2D([0],[0], color='#dc2626', linewidth=3.0, linestyle='--',
                   label=f'● 실패 평균선 (n={f_n})'),
        mpatches.Patch(color='#ef4444', alpha=0.2,
                       label='  실패 범위 (최솟값~최댓값)'),
        plt.Line2D([0],[0], color='#ef4444', linewidth=1.0, linestyle=':',
                   label='  ···  실패 최솟값/최댓값 경계'),
        plt.Line2D([0],[0], color='none',
                   label='▲ 각 축의 최댓값   ▼ 각 축의 최솟값'),
    ]
    ax.legend(handles=legend_handles, loc='lower left',
              bbox_to_anchor=(-0.25, -0.14), framealpha=0.25,
              facecolor=CARD, edgecolor=GRID, labelcolor=TEXT,
              fontsize=8.5, prop={'family': _FONT})

    # ── 컬러바 ────────────────────────────────────────────────
    ax_cb = fig.add_axes([0.04, 0.01, 0.50, 0.025])
    for xi in range(256):
        ax_cb.add_patch(plt.Rectangle((xi/256,0), 1/256, 1,
            color=prob_color(xi/255*100), transform=ax_cb.transAxes))
    ax_cb.set_xticks([0,.25,.5,.75,1.])
    ax_cb.set_xticklabels(['0%','25%','50%','75%','100%'],
                           color=TEXT, fontsize=8, fontfamily=_FONT)
    ax_cb.set_yticks([])
    ax_cb.set_title('성공 확률 색상 스케일', color=TEXT, fontsize=8, pad=4,
                    fontfamily=_FONT)
    for sp in ax_cb.spines.values(): sp.set_edgecolor(GRID)

    # ── 오른쪽 테이블 ─────────────────────────────────────────
    ax_t = fig.add_axes([0.60, 0.06, 0.39, 0.83])
    ax_t.set_facecolor(CARD); ax_t.axis('off')

    col_hdrs   = ["변수", "성공 평균", "실패 평균", "성공 확률"]
    col_ws     = [2.0, 1.1, 1.1, 1.2]
    total_w    = sum(col_ws)
    col_xs     = [0] + [sum(col_ws[:i+1])/total_w for i in range(len(col_ws)-1)]
    n_rows_tbl = len(var_names) + 2
    cell_h     = 0.88 / n_rows_tbl
    y_top      = 0.97

    def draw_cell(x, y, w, h, txt, bg, fg, bold=False, fs=13):
        ax_t.add_patch(mpatches.FancyBboxPatch(
            (x+.003, y+.003), w-.006, h-.006,
            boxstyle="round,pad=0.005", linewidth=0.5,
            edgecolor=GRID, facecolor=bg,
            transform=ax_t.transAxes, clip_on=False))
        ax_t.text(x+w/2, y+h/2, txt, ha='center', va='center',
                  fontsize=fs, color=fg,
                  fontweight='bold' if bold else 'normal',
                  transform=ax_t.transAxes, clip_on=False,
                  fontfamily=_FONT)

    # 헤더
    for ci, (hdr, cx, cw) in enumerate(zip(col_hdrs, col_xs, [w/total_w for w in col_ws])):
        draw_cell(cx, y_top-cell_h, cw, cell_h, hdr, "#1e2650", "#7eb3ff", bold=True, fs=12)

    # 데이터 행 (성공 확률 내림차순)
    sorted_vars = sorted(var_names, key=lambda v: -axis_probs[v]['prob'])
    row_bgs = ["#12152b","#161a30"]
    for ri, vname in enumerate(sorted_vars):
        info = axis_probs[vname]
        prob = info['prob']
        rgb  = prob_color(prob); hx = rgb_hex(rgb)
        y    = y_top - cell_h*(ri+2)
        cws  = [w/total_w for w in col_ws]

        draw_cell(col_xs[0], y, cws[0], cell_h, vname,  row_bgs[ri%2], TEXT)
        draw_cell(col_xs[1], y, cws[1], cell_h,
                  f"{info['s_mean']:.1f}", row_bgs[ri%2], '#4f8ef7')
        draw_cell(col_xs[2], y, cws[2], cell_h,
                  f"{info['f_mean']:.1f}", row_bgs[ri%2], '#ef4444')
        draw_cell(col_xs[3], y, cws[3], cell_h,
                  f"{prob}%", hx+'22', hx, bold=True, fs=14)

    # 전체
    y_ov = y_top - cell_h*(len(var_names)+2)
    ov_rgb = prob_color(base_rate); ov_hex = rgb_hex(ov_rgb)
    draw_cell(0, y_ov, 1.0, cell_h,
              f"기본 성공률 (전체): {base_rate:.1f}%",
              ov_hex+'22', ov_hex, bold=True, fs=15)

    ax_t.text(0.5, y_ov-0.025,
              "성공 확률 = 각 변수에서 성공 그룹이 실패 그룹보다\n높은 정도를 반영한 가중 추정값입니다.",
              ha='center', va='top', fontsize=11, color='#5c6490',
              transform=ax_t.transAxes, fontfamily=_FONT)

    plt.savefig(output_path, dpi=155, bbox_inches='tight',
                facecolor=BG, edgecolor='none')
    plt.close(fig)
    print(f"  [OK] 차트 저장: {output_path}")


# ─────────────────────────────────────────────────────────────
#  5b. Box Plot
# ─────────────────────────────────────────────────────────────
def draw_boxplot(df_vars: pd.DataFrame, labels: pd.Series,
                 axis_probs: dict, base_rate: float, output_path: Path):
    var_names = list(df_vars.columns)
    N = len(var_names)
    cols = 4
    rows = -(-N // cols)   # ceiling division

    fig, axes = plt.subplots(rows, cols, figsize=(22, rows * 3.8), facecolor=BG)
    fig.subplots_adjust(hspace=0.55, wspace=0.35, top=0.92, bottom=0.04,
                        left=0.05, right=0.98)
    fig.suptitle("변수별 성공 / 실패 분포 비교  (Box Plot)",
                 fontsize=16, color='white', fontweight='bold',
                 fontfamily=_FONT, y=0.97)

    axes_flat = axes.flatten() if rows > 1 else [axes] if cols == 1 else list(axes)

    for idx, vname in enumerate(var_names):
        ax = axes_flat[idx]
        ax.set_facecolor(CARD)
        ax.tick_params(colors=TEXT)
        for sp in ax.spines.values(): sp.set_edgecolor(GRID)

        s_vals = df_vars.loc[labels,  vname].dropna().values
        f_vals = df_vars.loc[~labels, vname].dropna().values

        bp = ax.boxplot([s_vals, f_vals],
                        patch_artist=True,
                        widths=0.45,
                        medianprops=dict(color='white', linewidth=2),
                        whiskerprops=dict(color=GRID, linewidth=1.2),
                        capprops=dict(color=GRID, linewidth=1.5),
                        flierprops=dict(marker='o', markersize=3,
                                        alpha=0.5, linestyle='none'))

        bp['boxes'][0].set_facecolor('#3b82f680')
        bp['boxes'][0].set_edgecolor('#60a5fa')
        bp['boxes'][1].set_facecolor('#ef444460')
        bp['boxes'][1].set_edgecolor('#f87171')

        prob = axis_probs[vname]['prob']
        rgb  = prob_color(prob)
        ax.set_title(f"{vname}  [{prob}%]", fontsize=9.5,
                     color=rgb_hex(rgb), fontweight='bold', fontfamily=_FONT)
        ax.set_xticks([1, 2])
        ax.set_xticklabels(['성공', '실패'], fontsize=9, fontfamily=_FONT, color=TEXT)
        ax.yaxis.set_tick_params(labelsize=8, labelcolor='#7080b8')
        ax.grid(axis='y', color=GRID, linestyle='--', alpha=0.5)
        ax.set_axisbelow(True)

    # 남은 칸 숨기기
    for idx in range(len(var_names), len(axes_flat)):
        axes_flat[idx].set_visible(False)

    plt.savefig(output_path, dpi=150, bbox_inches='tight',
                facecolor=BG, edgecolor='none')
    plt.close(fig)
    print(f"  [OK] 박스플롯 저장: {output_path}")


# ─────────────────────────────────────────────────────────────
#  5c. Heatmap
# ─────────────────────────────────────────────────────────────
def draw_heatmap(df_vars: pd.DataFrame, labels: pd.Series,
                 axis_probs: dict, base_rate: float, output_path: Path):
    var_names = list(df_vars.columns)

    # 정규화 (0~1)
    df_norm = df_vars.copy()
    for col in df_vars.columns:
        mn, mx = df_norm[col].min(), df_norm[col].max()
        df_norm[col] = (df_norm[col] - mn) / (mx - mn) if mx > mn else 0.5

    s_means = np.array([df_norm.loc[labels,  v].mean() for v in var_names])
    f_means = np.array([df_norm.loc[~labels, v].mean() for v in var_names])
    diff    = s_means - f_means   # 양수 = 성공이 높음

    # 성공 확률 기준 정렬
    order   = np.argsort([axis_probs[v]['prob'] for v in var_names])[::-1]
    var_sorted = [var_names[i] for i in order]
    s_sorted   = s_means[order]
    f_sorted   = f_means[order]
    d_sorted   = diff[order]

    data  = np.vstack([s_sorted, f_sorted, d_sorted]).T   # (N, 3)
    N     = len(var_names)

    fig, ax = plt.subplots(figsize=(10, max(6, N * 0.55 + 1.5)), facecolor=BG)
    fig.subplots_adjust(left=0.28, right=0.92, top=0.92, bottom=0.10)
    ax.set_facecolor(CARD)

    col_labels = ['성공 평균', '실패 평균', '차이 (성공-실패)']
    cmap_main  = plt.cm.Blues
    cmap_diff  = plt.cm.RdYlGn

    for ci, (col_lbl, cmap, vmin, vmax) in enumerate([
        ('성공 평균',        cmap_main, 0, 1),
        ('실패 평균',        cmap_main, 0, 1),
        ('차이 (성공-실패)', cmap_diff, -0.5, 0.5),
    ]):
        for ri in range(N):
            val  = data[ri, ci]
            norm = (val - vmin) / (vmax - vmin)
            norm = max(0, min(1, norm))
            color = cmap(norm)
            rect  = plt.Rectangle((ci, ri), 1, 1, facecolor=color,
                                   edgecolor='#1e2040', linewidth=0.5)
            ax.add_patch(rect)
            txt_val = f"{val:+.2f}" if ci == 2 else f"{val:.2f}"
            txt_col = 'white' if norm < 0.35 or norm > 0.75 else '#1a1a2e'
            ax.text(ci + 0.5, ri + 0.5, txt_val,
                    ha='center', va='center', fontsize=9.5,
                    color=txt_col, fontweight='bold', fontfamily=_FONT)

    ax.set_xlim(0, 3); ax.set_ylim(0, N)
    ax.set_xticks([0.5, 1.5, 2.5])
    ax.set_xticklabels(col_labels, fontsize=11, color=TEXT, fontfamily=_FONT)
    ax.set_yticks([i + 0.5 for i in range(N)])
    ax.set_yticklabels(var_sorted, fontsize=10, color=TEXT, fontfamily=_FONT)
    ax.tick_params(length=0)
    for sp in ax.spines.values(): sp.set_visible(False)

    # 성공 확률 색 막대 (우측)
    ax2 = ax.twinx()
    ax2.set_ylim(0, N)
    ax2.set_yticks([i + 0.5 for i in range(N)])
    probs = [axis_probs[v]['prob'] for v in var_sorted]
    ax2.set_yticklabels([f"{p}%" for p in probs],
                         fontsize=9.5, fontfamily=_FONT)
    ax2.tick_params(length=0)
    for ri, p in enumerate(probs):
        ax2.get_yticklabels()[ri].set_color(rgb_hex(prob_color(p)))
    for sp in ax2.spines.values(): sp.set_visible(False)
    ax2.set_ylabel("성공 확률", color='#7080b8', fontsize=10, fontfamily=_FONT)

    ax.set_title("변수별 성공/실패 히트맵  (정규화 0~1, 우측=성공확률)",
                 fontsize=13, color='white', fontweight='bold',
                 fontfamily=_FONT, pad=12)

    plt.savefig(output_path, dpi=150, bbox_inches='tight',
                facecolor=BG, edgecolor='none')
    plt.close(fig)
    print(f"  [OK] 히트맵 저장: {output_path}")


# ─────────────────────────────────────────────────────────────
#  5d. Parallel Coordinates
# ─────────────────────────────────────────────────────────────
def draw_parallel(df_vars: pd.DataFrame, labels: pd.Series,
                  axis_probs: dict, base_rate: float, output_path: Path):
    var_names = list(df_vars.columns)
    N = len(var_names)

    # 정규화
    df_norm = df_vars.copy()
    for col in df_vars.columns:
        mn, mx = df_norm[col].min(), df_norm[col].max()
        df_norm[col] = (df_norm[col] - mn) / (mx - mn) if mx > mn else 0.5

    fig, ax = plt.subplots(figsize=(22, 8), facecolor=BG)
    ax.set_facecolor(CARD)
    fig.subplots_adjust(top=0.88, bottom=0.18, left=0.03, right=0.97)

    # 케이스 라인 그리기 (실패 먼저, 성공 위에)
    for group, color, alpha, lw, zorder in [
        (False, '#ef4444', 0.18, 0.9, 3),
        (True,  '#3b82f6', 0.35, 1.1, 4),
    ]:
        mask = labels if group else ~labels
        for idx in df_norm[mask].index:
            row = [df_norm.loc[idx, v] for v in var_names]
            ax.plot(range(N), row, color=color, alpha=alpha,
                    linewidth=lw, zorder=zorder)

    # 성공/실패 평균선 (굵게)
    s_avg = [df_norm.loc[labels,  v].mean() for v in var_names]
    f_avg = [df_norm.loc[~labels, v].mean() for v in var_names]
    ax.plot(range(N), s_avg, color='#60a5fa', linewidth=3.5, zorder=6,
            marker='o', markersize=7, markeredgecolor='white', markeredgewidth=1.2)
    ax.plot(range(N), f_avg, color='#f87171', linewidth=3.5, zorder=6,
            linestyle='--', marker='s', markersize=7,
            markeredgecolor='white', markeredgewidth=1.2)

    # 축 라인 및 라벨
    for xi, vname in enumerate(var_names):
        ax.axvline(xi, color=GRID, linewidth=0.8, zorder=1)
        prob = axis_probs[vname]['prob']
        rgb  = prob_color(prob)
        ax.text(xi, -0.10, vname, ha='center', va='top', fontsize=8,
                color=TEXT, fontweight='bold', fontfamily=_FONT,
                transform=ax.get_xaxis_transform(),
                bbox=dict(boxstyle='round,pad=0.25', facecolor=rgb+(0.18,),
                          edgecolor=rgb+(0.5,), linewidth=0.6))
        ax.text(xi, -0.22, f"{prob}%", ha='center', va='top', fontsize=8.5,
                color=rgb_hex(rgb), fontweight='bold', fontfamily=_FONT,
                transform=ax.get_xaxis_transform())

    ax.set_xlim(-0.3, N - 0.7)
    ax.set_ylim(-0.05, 1.05)
    ax.set_yticks([0, 0.25, 0.5, 0.75, 1.0])
    ax.set_yticklabels(['최솟값', '25%', '50%', '75%', '최댓값'],
                        fontsize=9, color='#7080b8', fontfamily=_FONT)
    ax.set_xticks([])
    ax.grid(axis='y', color=GRID, linestyle='--', alpha=0.4)
    for sp in ax.spines.values(): sp.set_edgecolor(GRID)

    s_n = int(labels.sum()); f_n = int((~labels).sum())
    handles = [
        plt.Line2D([0],[0], color='#60a5fa', lw=3, label=f'성공 평균선 (n={s_n})'),
        plt.Line2D([0],[0], color='#3b82f6', lw=1.2, alpha=0.5, label='성공 케이스'),
        plt.Line2D([0],[0], color='#f87171', lw=3, ls='--', label=f'실패 평균선 (n={f_n})'),
        plt.Line2D([0],[0], color='#ef4444', lw=1.0, alpha=0.4, label='실패 케이스'),
    ]
    ax.legend(handles=handles, loc='upper right', framealpha=0.25,
              facecolor=CARD, edgecolor=GRID, labelcolor=TEXT,
              fontsize=9.5, prop={'family': _FONT})

    ax.set_title("평행좌표  (각 선 = 실험 케이스 1건, 굵은선 = 그룹 평균)",
                 fontsize=14, color='white', fontweight='bold',
                 fontfamily=_FONT, pad=10)

    plt.savefig(output_path, dpi=150, bbox_inches='tight',
                facecolor=BG, edgecolor='none')
    plt.close(fig)
    print(f"  [OK] 평행좌표 저장: {output_path}")


# ─────────────────────────────────────────────────────────────
#  5e. Lollipop Chart
# ─────────────────────────────────────────────────────────────
def draw_lollipop(df_vars: pd.DataFrame, labels: pd.Series,
                  axis_probs: dict, base_rate: float, output_path: Path):
    var_names = list(df_vars.columns)

    # 정규화
    df_norm = df_vars.copy()
    for col in df_vars.columns:
        mn, mx = df_norm[col].min(), df_norm[col].max()
        df_norm[col] = (df_norm[col] - mn) / (mx - mn) * 100 if mx > mn else 50.0

    s_means = {v: df_norm.loc[labels,  v].mean() for v in var_names}
    f_means = {v: df_norm.loc[~labels, v].mean() for v in var_names}
    diffs   = {v: s_means[v] - f_means[v] for v in var_names}

    # 차이값 기준 정렬
    sorted_vars = sorted(var_names, key=lambda v: diffs[v])
    N = len(sorted_vars)
    y = np.arange(N)

    fig, ax = plt.subplots(figsize=(14, max(7, N * 0.62 + 1.5)), facecolor=BG)
    ax.set_facecolor(CARD)
    fig.subplots_adjust(left=0.22, right=0.96, top=0.92, bottom=0.08)

    for i, vname in enumerate(sorted_vars):
        sm = s_means[vname]
        fm = f_means[vname]
        d  = diffs[vname]
        col = '#22c55e' if d >= 0 else '#ef4444'

        # 성공~실패 연결선
        ax.plot([fm, sm], [i, i], color=col, linewidth=2.2,
                alpha=0.7, zorder=2)
        # 실패 점
        ax.scatter(fm, i, color='#f87171', s=110, zorder=4,
                   edgecolors='white', linewidths=1.2)
        # 성공 점
        ax.scatter(sm, i, color='#60a5fa', s=110, zorder=4,
                   edgecolors='white', linewidths=1.2)
        # 차이 표시
        mid = (sm + fm) / 2
        ax.text(mid, i + 0.38, f"{d:+.1f}%", ha='center', va='bottom',
                fontsize=8.5, color=col, fontweight='bold', fontfamily=_FONT)

    # 성공 확률 색 라벨 (우측)
    for i, vname in enumerate(sorted_vars):
        prob = axis_probs[vname]['prob']
        ax.text(103, i, f"{prob}%", va='center', ha='left',
                fontsize=9, color=rgb_hex(prob_color(prob)),
                fontweight='bold', fontfamily=_FONT)

    ax.axvline(50, color=GRID, linewidth=1.0, linestyle=':', alpha=0.6)
    ax.set_xlim(0, 108)
    ax.set_ylim(-0.8, N - 0.2)
    ax.set_yticks(y)
    ax.set_yticklabels(sorted_vars, fontsize=10, fontfamily=_FONT, color=TEXT)
    ax.set_xlabel("정규화 값 (0~100%)", fontsize=10, color='#7080b8', fontfamily=_FONT)
    ax.xaxis.set_tick_params(labelsize=9, labelcolor='#7080b8')
    ax.grid(axis='x', color=GRID, linestyle='--', alpha=0.4)
    ax.set_axisbelow(True)
    for sp in ax.spines.values(): sp.set_edgecolor(GRID)

    s_n = int(labels.sum()); f_n = int((~labels).sum())
    handles = [
        plt.scatter([],[], color='#60a5fa', s=80, label=f'성공 평균 (n={s_n})'),
        plt.scatter([],[], color='#f87171', s=80, label=f'실패 평균 (n={f_n})'),
        plt.Line2D([0],[0], color='#22c55e', lw=2.5, label='연결선 (+: 성공이 높음)'),
        plt.Line2D([0],[0], color='#ef4444', lw=2.5, label='연결선 (-: 실패가 높음)'),
    ]
    ax.legend(handles=handles, loc='lower right', framealpha=0.25,
              facecolor=CARD, edgecolor=GRID, labelcolor=TEXT,
              fontsize=9.5, prop={'family': _FONT})

    ax.text(104, N - 0.5, "성공\n확률", ha='center', va='top',
            fontsize=8.5, color='#7080b8', fontfamily=_FONT)

    ax.set_title("변수별 성공 vs 실패 평균 비교  (Lollipop · 차이 기준 정렬)",
                 fontsize=14, color='white', fontweight='bold',
                 fontfamily=_FONT, pad=10)

    plt.savefig(output_path, dpi=150, bbox_inches='tight',
                facecolor=BG, edgecolor='none')
    plt.close(fig)
    print(f"  [OK] 롤리팝 저장: {output_path}")


# ─────────────────────────────────────────────────────────────
#  5f. Correlation Matrix Heatmap
# ─────────────────────────────────────────────────────────────
def draw_correlation(df_vars: pd.DataFrame, labels: pd.Series,
                     axis_probs: dict, base_rate: float, output_path: Path):
    """변수 간 피어슨 상관관계 행렬 히트맵 (세로 3단 배치, 큰 글자)"""
    var_names = list(df_vars.columns)
    # 변수명 축약 (괄호 안의 단위 제거해서 라벨 짧게)
    short_names = []
    for v in var_names:
        idx = v.find('(')
        short_names.append(v[:idx].strip() if idx > 0 else v)

    N = len(var_names)
    corr = df_vars.corr(method='pearson')

    # p-value 행렬 (t-검정 근사)
    n_obs = len(df_vars)
    pvals = pd.DataFrame(np.ones((N, N)), index=var_names, columns=var_names)
    for i in range(N):
        for j in range(i+1, N):
            r_val = float(corr.iloc[i, j])
            if abs(r_val) < 1.0 and n_obs > 2:
                t_stat = r_val * math.sqrt((n_obs - 2) / (1 - r_val**2 + 1e-12))
                p = 2 * math.exp(-0.5 * t_stat**2) / math.sqrt(2 * math.pi) if abs(t_stat) < 10 else 0
                p = max(0, min(1, p * math.sqrt(n_obs)))
            else:
                p = 0 if abs(r_val) >= 0.999 else 1
            pvals.iloc[i, j] = p
            pvals.iloc[j, i] = p

    # 성공/실패 그룹별 상관관계
    corr_s = df_vars[labels.values].corr()
    corr_f = df_vars[~labels.values].corr()

    # ── 세로 3단 배치 (컴팩트) ────────────────────────────
    cell_sz = max(0.45, min(0.7, 8 / N))
    mat_w = N * cell_sz
    fig_w = mat_w + 3
    fig_h = mat_w * 3 + 5

    fig = plt.figure(figsize=(fig_w, fig_h), facecolor=BG)

    fig.text(0.5, 0.995,
             "변수 간 상관관계 분석  (Pearson Correlation)",
             ha='center', fontsize=14, color='white', fontweight='bold',
             fontfamily=_FONT, va='top')

    cmap = plt.cm.RdBu_r

    matrices = [
        (corr,   "▣  전체 상관관계  (전체 데이터)", True),
        (corr_s, f"▣  성공 그룹 상관관계  (n={int(labels.sum())})", False),
        (corr_f, f"▣  실패 그룹 상관관계  (n={int((~labels).sum())})", False),
    ]

    # 각 행렬의 위치 (위에서 아래로)
    y_positions = [0.68, 0.36, 0.04]
    mat_height = 0.28

    for (mat, title, show_pval), y_base in zip(matrices, y_positions):
        ax = fig.add_axes([0.12, y_base, 0.78, mat_height])
        ax.set_facecolor(CARD)

        for i in range(N):
            for j in range(N):
                val = mat.iloc[i, j]
                norm_v = (val + 1) / 2

                # 대각선 = 짙은 회색
                if i == j:
                    fcolor = '#1a1e38'
                else:
                    fcolor = cmap(norm_v)

                rect_patch = plt.Rectangle((j, N - 1 - i), 1, 1,
                                            facecolor=fcolor,
                                            edgecolor='#0d0f1a',
                                            linewidth=1.2)
                ax.add_patch(rect_patch)

                if i == j:
                    ax.text(j + 0.5, N - 1 - i + 0.5, "1.00",
                            ha='center', va='center', fontsize=7,
                            color='#4a5280', fontfamily=_FONT)
                    continue

                txt = f"{val:.2f}"
                if show_pval:
                    p = pvals.iloc[i, j]
                    if p < 0.01:
                        txt += "**"
                    elif p < 0.05:
                        txt += "*"

                # 색상 대비
                if abs(val) > 0.5:
                    txt_col = 'white'
                    fw = 'bold'
                elif abs(val) > 0.25:
                    txt_col = '#e0e4f5'
                    fw = 'bold'
                else:
                    txt_col = '#8890b8'
                    fw = 'normal'

                ax.text(j + 0.5, N - 1 - i + 0.5, txt,
                        ha='center', va='center', fontsize=7,
                        color=txt_col, fontweight=fw, fontfamily=_FONT)

        ax.set_xlim(0, N)
        ax.set_ylim(0, N)
        ax.set_xticks([i + 0.5 for i in range(N)])
        ax.set_xticklabels(short_names, rotation=40, ha='right',
                           fontsize=7, color=TEXT, fontfamily=_FONT)
        ax.set_yticks([i + 0.5 for i in range(N)])
        ax.set_yticklabels(short_names[::-1], fontsize=7,
                           color=TEXT, fontfamily=_FONT)
        ax.tick_params(length=0, pad=4)
        for sp in ax.spines.values():
            sp.set_edgecolor(GRID)
            sp.set_linewidth(1)
        ax.set_title(title, fontsize=10, color='#7eb3ff', fontweight='bold',
                     fontfamily=_FONT, pad=8, loc='left')

    # 컬러바 (하단)
    ax_cb = fig.add_axes([0.20, 0.012, 0.60, 0.012])
    for xi in range(256):
        ax_cb.add_patch(plt.Rectangle((xi/256, 0), 1/256, 1,
            color=cmap(xi/255), transform=ax_cb.transAxes))
    ax_cb.set_xticks([0, 0.25, 0.5, 0.75, 1.0])
    ax_cb.set_xticklabels(['-1.0', '-0.5', '0', '+0.5', '+1.0'],
                           color=TEXT, fontsize=8, fontfamily=_FONT)
    ax_cb.set_yticks([])
    for sp in ax_cb.spines.values():
        sp.set_edgecolor(GRID)

    plt.savefig(output_path, dpi=130, bbox_inches='tight',
                facecolor=BG, edgecolor='none')
    plt.close(fig)
    print(f"  [OK] 상관관계 히트맵 저장: {output_path}")


# ─────────────────────────────────────────────────────────────
#  5g. Scatter Plot Matrix (주요 상관 변수 쌍)
# ─────────────────────────────────────────────────────────────
def draw_scatter_matrix(df_vars: pd.DataFrame, labels: pd.Series,
                        axis_probs: dict, base_rate: float, output_path: Path):
    """상관관계가 높은 상위 변수 쌍의 산점도 (성공/실패 색 구분, 큰 사이즈)"""
    var_names = list(df_vars.columns)
    N = len(var_names)

    # 변수명 축약
    def short(v):
        idx = v.find('(')
        return v[:idx].strip() if idx > 0 else v

    # 모든 변수 쌍의 상관계수
    pairs = []
    for i in range(N):
        for j in range(i+1, N):
            r = df_vars.iloc[:, i].corr(df_vars.iloc[:, j])
            pairs.append((abs(r), r, var_names[i], var_names[j]))
    pairs.sort(reverse=True)

    # 상위 9개 쌍 (3x3 그리드, 컴팩트)
    top_pairs = pairs[:min(9, len(pairs))]
    n_pairs = len(top_pairs)
    cols = min(3, n_pairs)
    rows = -(-n_pairs // cols)

    fig, axes = plt.subplots(rows, cols,
                              figsize=(cols * 4.5, rows * 3.8),
                              facecolor=BG)
    fig.subplots_adjust(hspace=0.50, wspace=0.35, top=0.92, bottom=0.04,
                        left=0.07, right=0.97)
    fig.suptitle("변수 간 상관관계 산점도  (상관계수 상위 쌍)",
                 fontsize=13, color='white', fontweight='bold',
                 fontfamily=_FONT, y=0.98)
    fig.text(0.5, 0.955, "● 성공  ● 실패  --- 추세선",
             ha='center', fontsize=9, color='#7080b8', fontfamily=_FONT)

    if rows == 1 and cols == 1:
        axes_flat = [axes]
    elif rows == 1:
        axes_flat = list(axes)
    else:
        axes_flat = axes.flatten()

    for idx, (abs_r, r, vx, vy) in enumerate(top_pairs):
        ax = axes_flat[idx]
        ax.set_facecolor('#0d1020')

        # 실패 먼저, 성공 위에
        ax.scatter(df_vars.loc[~labels, vx], df_vars.loc[~labels, vy],
                   c='#ef4444', alpha=0.45, s=25, edgecolors='#991b1b',
                   linewidths=0.5, zorder=3, label='실패')
        ax.scatter(df_vars.loc[labels, vx], df_vars.loc[labels, vy],
                   c='#3b82f6', alpha=0.6, s=25, edgecolors='#1d4ed8',
                   linewidths=0.5, zorder=4, label='성공')

        # 추세선
        mask = df_vars[[vx, vy]].dropna().index
        if len(mask) > 2:
            z = np.polyfit(df_vars.loc[mask, vx], df_vars.loc[mask, vy], 1)
            xline = np.linspace(df_vars[vx].min(), df_vars[vx].max(), 50)
            ax.plot(xline, np.polyval(z, xline), color='#fbbf24',
                    linewidth=2.2, linestyle='--', alpha=0.8, zorder=5)

        # 상관 강도별 색상 + 배지
        if abs_r > 0.5:
            cor_color = '#22c55e'
            badge = "강한 상관"
        elif abs_r > 0.3:
            cor_color = '#fbbf24'
            badge = "보통 상관"
        else:
            cor_color = '#6b7280'
            badge = "약한 상관"

        ax.set_title(f"{short(vx)}  vs  {short(vy)}",
                     fontsize=9, color='white', fontweight='bold',
                     fontfamily=_FONT, pad=6)
        # r 값 배지 (차트 안 좌상단)
        ax.text(0.04, 0.96, f"r={r:+.3f} ({badge})",
                transform=ax.transAxes, fontsize=7.5, color=cor_color,
                fontweight='bold', fontfamily=_FONT, va='top',
                bbox=dict(boxstyle='round,pad=0.3', facecolor='#0d0f1a',
                          edgecolor=cor_color, alpha=0.9, linewidth=0.8))

        ax.set_xlabel(short(vx), fontsize=8, color='#7080b8',
                      fontfamily=_FONT, labelpad=4)
        ax.set_ylabel(short(vy), fontsize=8, color='#7080b8',
                      fontfamily=_FONT, labelpad=4)
        ax.tick_params(colors='#5c6490', labelsize=7)
        ax.grid(color=GRID, linestyle='--', alpha=0.3)
        for sp in ax.spines.values():
            sp.set_edgecolor(GRID)
            sp.set_linewidth(1.2)

        if idx == 0:
            ax.legend(loc='lower right', framealpha=0.4,
                      facecolor='#0d0f1a', edgecolor=GRID, labelcolor=TEXT,
                      fontsize=7, prop={'family': _FONT},
                      markerscale=1.0)

    # 남은 칸 숨기기
    for idx in range(n_pairs, len(axes_flat)):
        axes_flat[idx].set_visible(False)

    plt.savefig(output_path, dpi=130, bbox_inches='tight',
                facecolor=BG, edgecolor='none')
    plt.close(fig)
    print(f"  [OK] 산점도 매트릭스 저장: {output_path}")


# ─────────────────────────────────────────────────────────────
#  6. HTML 보고서 (탭 구조 · 5개 차트 + 예측 계산기)
# ─────────────────────────────────────────────────────────────
def generate_html_report(df_vars, labels, axis_probs, base_rate,
                         charts: dict, output_path: Path, source_file: str):

    var_names = list(df_vars.columns)

    def pcolor(p):
        t = p/100.0
        if t < 0.5:
            r=int(239+(251-239)*t*2); g=int(68+(159-68)*t*2); b=68
        else:
            t2=(t-0.5)*2
            r=int(251-(251-34)*t2); g=int(159+(197-159)*t2); b=int(68-(68-94)*t2)
        return f"#{r:02x}{g:02x}{b:02x}"

    s_n = int(labels.sum()); f_n = int((~labels).sum()); tot = len(labels)

    # 테이블 행 (성공 확률 내림차순)
    sorted_vars = sorted(var_names, key=lambda v: -axis_probs[v]['prob'])
    rows_html = ""
    for ri, vname in enumerate(sorted_vars):
        info = axis_probs[vname]
        prob = info['prob']
        col  = pcolor(prob)
        rows_html += f"""
        <tr class="{'odd' if ri%2==0 else 'even'}">
          <td class="vname">{vname}</td>
          <td style="color:#4f8ef7;font-weight:600">{info['s_mean']:.2f}</td>
          <td style="color:#ef4444;font-weight:600">{info['f_mean']:.2f}</td>
          <td style="color:#888">{info['col_min']:.1f} ~ {info['col_max']:.1f}</td>
          <td>
            <div class="pbar-wrap">
              <div class="pbar" style="width:{prob}%;background:linear-gradient(to right,{pcolor(max(0,prob-25))},{col})"></div>
              <span style="color:{col};font-weight:700;font-family:'JetBrains Mono'">{prob}%</span>
            </div>
          </td>
        </tr>"""

    # 상관관계 행렬 데이터 (JS용)
    corr_matrix = df_vars.corr(method='pearson')
    js_corr = []
    for i, vi in enumerate(var_names):
        row_corr = []
        for j, vj in enumerate(var_names):
            row_corr.append(round(float(corr_matrix.loc[vi, vj]), 4))
        js_corr.append(row_corr)

    # 상관관계 테이블 HTML (상위 쌍 + 전체)
    corr_pairs = []
    for i in range(len(var_names)):
        for j in range(i+1, len(var_names)):
            r = float(corr_matrix.iloc[i, j])
            corr_pairs.append((abs(r), r, var_names[i], var_names[j]))
    corr_pairs.sort(reverse=True)

    corr_rows_html = ""
    for ri, (abs_r, r, vx, vy) in enumerate(corr_pairs):
        if abs_r > 0.5:
            strength = "강한"
            badge_col = "#22c55e"
        elif abs_r > 0.3:
            strength = "보통"
            badge_col = "#fbbf24"
        else:
            strength = "약한"
            badge_col = "#6b7280"
        direction = "양의 상관" if r > 0 else "음의 상관"
        bar_w = int(abs_r * 100)
        bar_col = "#3b82f6" if r > 0 else "#ef4444"
        corr_rows_html += f"""
        <tr class="{'odd' if ri%2==0 else 'even'}" style="cursor:pointer"
            onclick="setScatterVars('{vx}','{vy}')">
          <td class="vname">{vx}</td>
          <td class="vname">{vy}</td>
          <td style="font-family:'JetBrains Mono';font-weight:700;color:{bar_col}">{r:+.4f}</td>
          <td>
            <div class="pbar-wrap">
              <div class="pbar" style="width:{bar_w}%;background:{bar_col}"></div>
              <span style="color:{badge_col};font-weight:600;font-size:11px">{strength} {direction}</span>
            </div>
          </td>
        </tr>"""

    # 예측 입력 필드 JS 데이터
    js_vars = str(var_names).replace("'", '"')
    js_mins = str([axis_probs[v]['col_min'] for v in var_names])
    js_maxs = str([axis_probs[v]['col_max'] for v in var_names])
    js_s_means = str([round(axis_probs[v]['s_mean'], 2) for v in var_names])

    # 모든 케이스 데이터 (JS용)
    js_cases  = []
    js_labels_arr = []
    for idx, row in df_vars.iterrows():
        js_cases.append([float(row[c]) for c in var_names])
        js_labels_arr.append(1 if labels.iloc[idx] else 0)

    oc = pcolor(base_rate)

    # 예측 입력 폼
    input_fields = ""
    for vname in var_names:
        info = axis_probs[vname]
        mn, mx = info['col_min'], info['col_max']
        mid    = round((mn + mx) / 2, 1)
        input_fields += f"""
        <div class="pred-field">
          <label>{vname}<span class="range-hint">{mn:.0f}~{mx:.0f}</span></label>
          <input type="number" id="pred_{vname.replace(' ','_')}"
                 min="{mn}" max="{mx}" step="1" value="{mid}"
                 onchange="calcPred()" oninput="calcPred()"/>
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<title>방사형 분석 보고서 · 성공/실패</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet"/>
<style>
:root{{--bg:#0d0f1a;--card:#12152b;--card2:#161a30;--border:rgba(79,142,247,.15);
      --text:#e0e4f5;--muted:#6b7498;--acc:#4f8ef7;}}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Inter',sans-serif;background:var(--bg);color:var(--text);
     min-height:100vh;padding:28px 24px;max-width:1400px;margin:0 auto;
     font-size:15px;line-height:1.6}}
h1{{font-size:28px;font-weight:700;
    background:linear-gradient(135deg,#4f8ef7,#a78bfa);
    -webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent;
    margin-bottom:6px}}
.meta{{font-size:13px;color:var(--muted);margin-bottom:28px}}
/* stat cards */
.stat-row{{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:28px}}
.stat-card{{background:var(--card);border:1px solid var(--border);
           border-radius:14px;padding:20px 24px;text-align:center}}
.stat-num{{font-size:38px;font-weight:700;font-family:'JetBrains Mono'}}
.stat-lbl{{font-size:12px;color:var(--muted);text-transform:uppercase;
          letter-spacing:.8px;margin-top:6px}}
/* chart */
img.chart{{width:100%;border-radius:14px;border:1px solid var(--border);
          margin-bottom:24px}}
/* card */
.card{{background:var(--card);border:1px solid var(--border);
      border-radius:16px;padding:28px;margin-bottom:24px}}
.sec-title{{font-size:14px;font-weight:700;color:#8b95c8;text-transform:uppercase;
           letter-spacing:1.2px;margin-bottom:18px;
           display:flex;align-items:center;gap:10px}}
.sec-title::before{{content:'';width:4px;height:18px;
                   background:linear-gradient(to bottom,#4f8ef7,#a78bfa);border-radius:2px}}
/* table */
table{{width:100%;border-collapse:collapse;font-size:14px}}
th{{background:#1a1e38;padding:12px 16px;text-align:left;
   font-size:12px;text-transform:uppercase;letter-spacing:.6px;
   color:var(--muted);border-bottom:2px solid var(--border);
   position:sticky;top:0;z-index:1}}
td{{padding:12px 16px;border-bottom:1px solid rgba(255,255,255,.05)}}
tr.odd td{{background:rgba(255,255,255,.015)}}
tr:hover td{{background:rgba(79,142,247,.07);transition:background .15s}}
.vname{{font-weight:600;color:var(--text);white-space:nowrap}}
.pbar-wrap{{display:flex;align-items:center;gap:10px}}
.pbar{{height:10px;border-radius:5px;min-width:4px;transition:width .5s}}
/* prediction */
.pred-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(190px,1fr));gap:16px;
           margin-bottom:22px}}
.pred-field label{{display:block;font-size:12px;font-weight:600;color:var(--muted);
                  text-transform:uppercase;letter-spacing:.6px;margin-bottom:8px}}
.range-hint{{font-size:11px;color:#4a5280;margin-left:6px;font-weight:400}}
.pred-field input,.pred-field select{{width:100%;background:#080a16;
                  border:1px solid rgba(79,142,247,.2);
                  border-radius:10px;padding:11px 14px;color:var(--text);
                  font-size:15px;font-family:'JetBrains Mono';
                  transition:border-color .2s,box-shadow .2s}}
.pred-field input:focus,.pred-field select:focus{{outline:none;border-color:var(--acc);
                         box-shadow:0 0 0 4px rgba(79,142,247,.15)}}
.pred-field select{{cursor:pointer;appearance:none;
                   background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 12 12'%3E%3Cpath d='M2 4l4 4 4-4' fill='none' stroke='%236b7498' stroke-width='2'/%3E%3C/svg%3E");
                   background-repeat:no-repeat;background-position:right 12px center;
                   padding-right:36px}}
.pred-result{{text-align:center;padding:28px;
             background:linear-gradient(135deg,rgba(79,142,247,.06),rgba(167,139,250,.06));
             border:1px solid rgba(79,142,247,.18);border-radius:14px}}
.pred-num{{font-size:64px;font-weight:700;font-family:'JetBrains Mono';
          line-height:1;transition:color .4s}}
.pred-desc{{font-size:14px;color:var(--muted);margin-top:10px}}
.pred-bar-bg{{height:16px;border-radius:8px;background:rgba(255,255,255,.06);
             margin-top:16px;overflow:hidden}}
.pred-bar-fill{{height:100%;border-radius:8px;transition:width .6s,background .4s}}
.calc-btn{{width:100%;padding:14px;margin-top:18px;
          background:linear-gradient(135deg,#4f8ef7,#a78bfa);
          border:none;border-radius:12px;color:white;
          font-size:16px;font-weight:600;cursor:pointer;
          font-family:'Inter',sans-serif;transition:all .3s}}
.calc-btn:hover{{transform:translateY(-2px);
                box-shadow:0 12px 32px rgba(79,142,247,.35)}}
/* tabs */
.tab-bar{{display:flex;background:#0a0c16;border-bottom:2px solid #1e2440;
          overflow-x:auto;-webkit-overflow-scrolling:touch}}
.tab-btn{{flex:0 0 auto;padding:14px 18px;background:none;border:none;
          color:var(--muted);font-size:14px;font-weight:600;white-space:nowrap;
          cursor:pointer;transition:all .2s;font-family:'Inter',sans-serif;
          border-bottom:3px solid transparent;margin-bottom:-2px}}
.tab-btn.active{{color:#7eb3ff;border-bottom-color:#4f8ef7;
                background:rgba(79,142,247,.05)}}
.tab-btn:hover{{color:white;background:rgba(255,255,255,.03)}}
.tab-btn.corr-tab{{color:#fbbf24}}
.tab-btn.corr-tab.active{{color:#fde68a;border-bottom-color:#fbbf24}}
.tab-pane{{display:none}}
.tab-pane.active{{display:block}}
/* scatter explorer */
.scatter-wrap{{display:grid;grid-template-columns:1fr 380px;gap:20px;align-items:start}}
.scatter-stats-card{{background:#0a0c18;border:1px solid var(--border);
                    border-radius:14px;padding:20px;font-size:13px}}
.scatter-stat-row{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px}}
.scatter-stat-box{{padding:12px;border-radius:10px;text-align:center}}
.scatter-stat-box .label{{font-size:11px;font-weight:600;margin-bottom:4px}}
.scatter-stat-box .value{{font-size:20px;font-weight:700;font-family:'JetBrains Mono'}}
/* corr table clickable rows */
table.corr-table tr{{cursor:pointer;transition:background .15s}}
table.corr-table tr:hover td{{background:rgba(251,191,36,.08)}}
footer{{text-align:center;color:var(--muted);font-size:12px;
       margin-top:36px;padding-top:18px;border-top:1px solid var(--border)}}
/* responsive for smaller screens / notebook */
@media (max-width:900px) {{
  body{{padding:16px 12px;font-size:14px}}
  .stat-row{{grid-template-columns:repeat(2,1fr)}}
  .scatter-wrap{{grid-template-columns:1fr}}
  .pred-grid{{grid-template-columns:repeat(auto-fill,minmax(140px,1fr))}}
  .tab-btn{{padding:10px 12px;font-size:12px}}
}}
</style>
<script>
function showTab(id,btn){{
  document.querySelectorAll('.tab-pane').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById('tab-'+id).classList.add('active');
  btn.classList.add('active');
}}
</script>
</head>
<body>
<h1>방사형 데이터 분석 보고서</h1>
<p class="meta">소스: {source_file} | 생성: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')} |
   변수 {len(var_names)}개 · 전체 {tot}건</p>

<div class="stat-row">
  <div class="stat-card">
    <div class="stat-num" style="color:#e0e4f5">{tot}</div>
    <div class="stat-lbl">전체 케이스</div>
  </div>
  <div class="stat-card">
    <div class="stat-num" style="color:#4f8ef7">{s_n}</div>
    <div class="stat-lbl">성공</div>
  </div>
  <div class="stat-card">
    <div class="stat-num" style="color:#ef4444">{f_n}</div>
    <div class="stat-lbl">실패</div>
  </div>
  <div class="stat-card">
    <div class="stat-num" style="color:{oc}">{base_rate:.1f}%</div>
    <div class="stat-lbl">기본 성공률</div>
  </div>
</div>

<!-- 차트 탭 -->
<div class="card" style="padding:0;overflow:hidden;border-radius:16px">
  <div class="tab-bar">
    <button class="tab-btn active" onclick="showTab('corr',this)">① 상관관계 히트맵</button>
    <button class="tab-btn"        onclick="showTab('scatter',this)">② 산점도 매트릭스</button>
  </div>
  <div id="tab-corr"    class="tab-pane active" style="padding:8px"><img src="{charts['corr'].name}"    style="width:100%;border-radius:10px"></div>
  <div id="tab-scatter" class="tab-pane"        style="padding:8px"><img src="{charts['scatter'].name}" style="width:100%;border-radius:10px"></div>
</div>

<!-- 예측 계산기 -->
<div class="card">
  <div class="sec-title">예측값 입력 → 성공 확률 계산</div>
  <p style="font-size:13px;color:var(--muted);margin-bottom:18px">
    각 변수에 예측값을 입력하면 AI가 실제 데이터를 기반으로 성공 확률을 계산합니다.</p>

  <div class="pred-grid">
    {input_fields}
  </div>

  <button class="calc-btn" onclick="calcPred()">계산하기</button>

  <div class="pred-result" style="margin-top:20px">
    <div style="font-size:12px;color:var(--muted);text-transform:uppercase;letter-spacing:1px">예측 성공 확률</div>
    <div class="pred-num" id="predNum">—</div>
    <div class="pred-desc" id="predDesc">위 값을 입력하고 계산하기를 누르세요</div>
    <div class="pred-bar-bg">
      <div class="pred-bar-fill" id="predBar" style="width:0%;background:#4f8ef7"></div>
    </div>
  </div>
</div>

<!-- 인터랙티브 상관관계 탐색기 -->
<div class="card">
  <div class="sec-title">변수 간 상관관계 탐색기</div>
  <p style="font-size:14px;color:var(--muted);margin-bottom:22px;line-height:1.6">
    두 변수를 선택하면 산점도와 상관 통계를 실시간으로 확인할 수 있습니다.<br>
    <span style="color:#fbbf24">아래 상관계수 테이블의 행을 클릭</span>해도 자동 선택됩니다.</p>

  <div style="display:grid;grid-template-columns:1fr 1fr 200px;gap:18px;margin-bottom:22px;align-items:end">
    <div class="pred-field">
      <label>X축 변수</label>
      <select id="scatterX" onchange="drawScatter()">
        {"".join(f'<option value="{i}">{v}</option>' for i, v in enumerate(var_names))}
      </select>
    </div>
    <div class="pred-field">
      <label>Y축 변수</label>
      <select id="scatterY" onchange="drawScatter()">
        {"".join(f'<option value="{i}" {"selected" if i==1 else ""}>{v}</option>' for i, v in enumerate(var_names))}
      </select>
    </div>
    <div style="text-align:center;padding-bottom:4px">
      <div style="font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:1px;margin-bottom:4px">상관계수</div>
      <div id="scatterR" style="font-size:46px;font-weight:700;font-family:'JetBrains Mono';color:#4f8ef7;line-height:1">—</div>
    </div>
  </div>

  <div class="scatter-wrap">
    <div>
      <canvas id="scatterCanvas" width="700" height="500"
              style="width:100%;border:1px solid var(--border);border-radius:14px;
                     background:#080a14"></canvas>
    </div>
    <div id="scatterStats" style="font-size:13px;color:var(--text)">
    </div>
  </div>
</div>

<!-- 상관관계 테이블 (전체 변수 쌍) -->
<div class="card">
  <div class="sec-title">변수 간 상관계수 테이블</div>
  <p style="font-size:13px;color:var(--muted);margin-bottom:16px">
    행을 클릭하면 위 산점도에 해당 변수 쌍이 표시됩니다. 상관계수 절댓값 기준 내림차순.</p>
  <div style="max-height:500px;overflow-y:auto;border-radius:10px;border:1px solid var(--border)">
  <table class="corr-table">
    <thead>
      <tr>
        <th style="width:25%">변수 X</th>
        <th style="width:25%">변수 Y</th>
        <th style="width:18%">상관계수 (r)</th>
        <th>강도</th>
      </tr>
    </thead>
    <tbody>{corr_rows_html}</tbody>
  </table>
  </div>
</div>

<!-- 축별 성공 확률 테이블 -->
<div class="card">
  <div class="sec-title">변수별 성공 확률 분석</div>
  <table>
    <thead>
      <tr>
        <th>변수</th>
        <th style="color:#4f8ef7">성공 평균</th>
        <th style="color:#ef4444">실패 평균</th>
        <th>범위</th>
        <th>성공 확률</th>
      </tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
</div>

<footer>Radar Analysis v2 · 성공/실패 + 상관관계 분석 | 예측: KNN 가중 거리 방식 (k=7)</footer>

<script>
// 케이스 데이터 (KNN 계산용)
const VARS    = {js_vars};
const MINS    = {js_mins};
const MAXS    = {js_maxs};
const CASES   = {str(js_cases).replace("'",'"')};
const LABELSC = {str(js_labels_arr).replace("'",'"')};
const CORR    = {str(js_corr)};

function getInputVals() {{
  return VARS.map(v => {{
    const el = document.getElementById('pred_' + v.replace(/ /g,'_'));
    return el ? parseFloat(el.value) || 0 : 0;
  }});
}}

function knnPredict(inputVals, k=7) {{
  // 정규화
  const norm = inputVals.map((v,i) => {{
    const rng = MAXS[i] - MINS[i];
    return rng > 0 ? (v - MINS[i]) / rng : 0.5;
  }});
  // 거리 계산
  const dists = CASES.map(c => {{
    const normC = c.map((v,i) => {{
      const rng = MAXS[i] - MINS[i];
      return rng > 0 ? (v - MINS[i]) / rng : 0.5;
    }});
    return Math.sqrt(normC.reduce((s,v,i) => s + (v-norm[i])**2, 0)) + 1e-9;
  }});
  // 상위 k
  const idx = dists.map((d,i)=>[d,i]).sort((a,b)=>a[0]-b[0]).slice(0,k);
  const wsum = idx.reduce((s,[d])=>s+1/d, 0);
  const prob = idx.reduce((s,[d,i])=>s+(1/d)*LABELSC[i], 0) / wsum;
  return Math.round(prob * 1000) / 10;
}}

function colorForProb(p) {{
  const t = p/100;
  if(t < 0.5) {{
    const r=Math.round(239+(251-239)*t*2);
    const g=Math.round(68+(159-68)*t*2);
    return `rgb(${{r}},${{g}},68)`;
  }} else {{
    const t2=(t-0.5)*2;
    const r=Math.round(251-(251-34)*t2);
    const g=Math.round(159+(197-159)*t2);
    const b=Math.round(68+(94-68)*t2);
    return `rgb(${{r}},${{g}},${{b}})`;
  }}
}}

function calcPred() {{
  const vals = getInputVals();
  const prob = knnPredict(vals, 7);
  const col  = colorForProb(prob);
  document.getElementById('predNum').textContent  = prob + '%';
  document.getElementById('predNum').style.color  = col;
  document.getElementById('predBar').style.width  = prob + '%';
  document.getElementById('predBar').style.background =
    `linear-gradient(to right, ${{colorForProb(Math.max(0,prob-25))}}, ${{col}})`;

  let desc = '';
  if(prob >= 70)      desc = '성공 가능성이 높습니다! 유사한 과거 케이스의 대부분이 성공했습니다.';
  else if(prob >= 45) desc = '성공/실패 경계선에 있습니다. 핵심 변수를 개선하면 확률이 올라갑니다.';
  else                desc = '성공 가능성이 낮습니다. 유사한 과거 케이스의 대부분이 실패했습니다.';
  document.getElementById('predDesc').textContent = desc;
}}

// ── 인터랙티브 산점도 ──────────────────────────────
function setScatterVars(vx, vy) {{
  const xi = VARS.indexOf(vx);
  const yi = VARS.indexOf(vy);
  if (xi >= 0 && yi >= 0) {{
    document.getElementById('scatterX').value = xi;
    document.getElementById('scatterY').value = yi;
    drawScatter();
    // 스크롤 이동
    document.getElementById('scatterCanvas').scrollIntoView({{behavior:'smooth',block:'center'}});
  }}
}}

function drawScatter() {{
  const xi = parseInt(document.getElementById('scatterX').value);
  const yi = parseInt(document.getElementById('scatterY').value);
  const canvas = document.getElementById('scatterCanvas');
  const ctx = canvas.getContext('2d');

  // 고해상도 지원
  const dpr = window.devicePixelRatio || 1;
  const w = canvas.clientWidth;
  const h = canvas.clientHeight;
  canvas.width = w * dpr;
  canvas.height = h * dpr;
  ctx.scale(dpr, dpr);

  const pad = {{top:30, right:30, bottom:50, left:60}};
  const pw = w - pad.left - pad.right;
  const ph = h - pad.top - pad.bottom;

  // 데이터 추출
  const xData = CASES.map(c => c[xi]);
  const yData = CASES.map(c => c[yi]);
  const xMin = Math.min(...xData), xMax = Math.max(...xData);
  const yMin = Math.min(...yData), yMax = Math.max(...yData);
  const xRng = xMax - xMin || 1, yRng = yMax - yMin || 1;

  // 배경
  ctx.fillStyle = '#0a0c18';
  ctx.fillRect(0, 0, w, h);

  // 그리드
  ctx.strokeStyle = '#1e2240';
  ctx.lineWidth = 0.5;
  for (let i = 0; i <= 5; i++) {{
    const gx = pad.left + pw * i / 5;
    const gy = pad.top + ph * i / 5;
    ctx.beginPath(); ctx.moveTo(gx, pad.top); ctx.lineTo(gx, pad.top + ph); ctx.stroke();
    ctx.beginPath(); ctx.moveTo(pad.left, gy); ctx.lineTo(pad.left + pw, gy); ctx.stroke();

    ctx.fillStyle = '#5c6490'; ctx.font = '10px JetBrains Mono';
    ctx.textAlign = 'center';
    ctx.fillText((xMin + xRng * i / 5).toFixed(1), gx, h - pad.bottom + 16);
    ctx.textAlign = 'right';
    ctx.fillText((yMax - yRng * i / 5).toFixed(1), pad.left - 8, gy + 4);
  }}

  // 축 라벨
  ctx.fillStyle = '#7080b8'; ctx.font = '12px Inter';
  ctx.textAlign = 'center';
  ctx.fillText(VARS[xi], pad.left + pw / 2, h - 6);
  ctx.save();
  ctx.translate(14, pad.top + ph / 2);
  ctx.rotate(-Math.PI / 2);
  ctx.fillText(VARS[yi], 0, 0);
  ctx.restore();

  // 좌표 변환
  function tx(v) {{ return pad.left + (v - xMin) / xRng * pw; }}
  function ty(v) {{ return pad.top + ph - (v - yMin) / yRng * ph; }}

  // 실패 점 먼저
  for (let i = 0; i < CASES.length; i++) {{
    if (LABELSC[i] === 1) continue;
    ctx.beginPath();
    ctx.arc(tx(xData[i]), ty(yData[i]), 4.5, 0, Math.PI * 2);
    ctx.fillStyle = 'rgba(239,68,68,0.45)';
    ctx.fill();
    ctx.strokeStyle = 'rgba(153,27,27,0.6)';
    ctx.lineWidth = 0.8;
    ctx.stroke();
  }}
  // 성공 점
  for (let i = 0; i < CASES.length; i++) {{
    if (LABELSC[i] === 0) continue;
    ctx.beginPath();
    ctx.arc(tx(xData[i]), ty(yData[i]), 4.5, 0, Math.PI * 2);
    ctx.fillStyle = 'rgba(59,130,246,0.55)';
    ctx.fill();
    ctx.strokeStyle = 'rgba(29,78,216,0.6)';
    ctx.lineWidth = 0.8;
    ctx.stroke();
  }}

  // 추세선
  const n = xData.length;
  const sx = xData.reduce((a,b) => a+b, 0);
  const sy = yData.reduce((a,b) => a+b, 0);
  const sxy = xData.reduce((a, v, i) => a + v * yData[i], 0);
  const sx2 = xData.reduce((a, v) => a + v * v, 0);
  const slope = (n * sxy - sx * sy) / (n * sx2 - sx * sx);
  const intercept = (sy - slope * sx) / n;

  ctx.beginPath();
  ctx.moveTo(tx(xMin), ty(slope * xMin + intercept));
  ctx.lineTo(tx(xMax), ty(slope * xMax + intercept));
  ctx.strokeStyle = 'rgba(251,191,36,0.7)';
  ctx.lineWidth = 2;
  ctx.setLineDash([6, 4]);
  ctx.stroke();
  ctx.setLineDash([]);

  // 범례
  ctx.font = '11px Inter';
  const lx = pad.left + pw - 130, ly = pad.top + 12;
  ctx.fillStyle = 'rgba(18,21,43,0.85)';
  ctx.fillRect(lx - 8, ly - 6, 140, 52);
  ctx.fillStyle = 'rgba(59,130,246,0.7)';
  ctx.beginPath(); ctx.arc(lx, ly + 6, 4, 0, Math.PI*2); ctx.fill();
  ctx.fillStyle = '#c0c8f0'; ctx.textAlign = 'left';
  ctx.fillText('성공', lx + 10, ly + 10);
  ctx.fillStyle = 'rgba(239,68,68,0.6)';
  ctx.beginPath(); ctx.arc(lx, ly + 24, 4, 0, Math.PI*2); ctx.fill();
  ctx.fillStyle = '#c0c8f0';
  ctx.fillText('실패', lx + 10, ly + 28);
  ctx.strokeStyle = 'rgba(251,191,36,0.7)'; ctx.lineWidth = 2;
  ctx.setLineDash([4,3]);
  ctx.beginPath(); ctx.moveTo(lx - 5, ly + 40); ctx.lineTo(lx + 5, ly + 40); ctx.stroke();
  ctx.setLineDash([]);
  ctx.fillStyle = '#c0c8f0';
  ctx.fillText('추세선', lx + 10, ly + 44);

  // 상관계수 표시
  const r = CORR[xi][yi];
  const rEl = document.getElementById('scatterR');
  rEl.textContent = r.toFixed(4);
  rEl.style.color = Math.abs(r) > 0.5 ? '#22c55e' : (Math.abs(r) > 0.3 ? '#fbbf24' : '#6b7280');

  // 통계 패널
  const sXdata = xData.filter((_,i) => LABELSC[i]===1);
  const fXdata = xData.filter((_,i) => LABELSC[i]===0);
  const sYdata = yData.filter((_,i) => LABELSC[i]===1);
  const fYdata = yData.filter((_,i) => LABELSC[i]===0);
  const avg = arr => arr.length ? (arr.reduce((a,b)=>a+b,0)/arr.length) : 0;
  const std = arr => {{
    if(arr.length < 2) return 0;
    const m = avg(arr);
    return Math.sqrt(arr.reduce((s,v) => s + (v-m)**2, 0) / (arr.length-1));
  }};

  // 성공/실패 그룹별 상관계수
  function pearson(xs, ys) {{
    const n = xs.length;
    if (n < 3) return 0;
    const mx = avg(xs), my = avg(ys);
    let num = 0, dx2 = 0, dy2 = 0;
    for (let i = 0; i < n; i++) {{
      const dx = xs[i] - mx, dy = ys[i] - my;
      num += dx * dy; dx2 += dx*dx; dy2 += dy*dy;
    }}
    return dx2 > 0 && dy2 > 0 ? num / Math.sqrt(dx2 * dy2) : 0;
  }}

  const rS = pearson(sXdata, sYdata);
  const rF = pearson(fXdata, fYdata);

  const strength = Math.abs(r) > 0.7 ? '매우 강한' :
                   Math.abs(r) > 0.5 ? '강한' :
                   Math.abs(r) > 0.3 ? '보통' :
                   Math.abs(r) > 0.1 ? '약한' : '거의 없는';
  const dir = r > 0 ? '양의 상관' : r < 0 ? '음의 상관' : '무상관';

  document.getElementById('scatterStats').innerHTML = `
    <div style="background:var(--card);border:1px solid var(--border);border-radius:12px;padding:18px;margin-bottom:14px">
      <div style="font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:1px;margin-bottom:10px">상관 분석 결과</div>
      <div style="font-size:15px;font-weight:600;margin-bottom:12px;color:${{Math.abs(r)>0.5?'#22c55e':(Math.abs(r)>0.3?'#fbbf24':'#6b7280')}}">
        ${{strength}} ${{dir}} (r = ${{r.toFixed(4)}})
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;font-size:12px">
        <div style="padding:8px;background:rgba(59,130,246,.08);border-radius:8px">
          <div style="color:#60a5fa;font-weight:600">성공 그룹 r</div>
          <div style="font-family:'JetBrains Mono';font-size:16px;font-weight:700;color:#4f8ef7">${{rS.toFixed(4)}}</div>
        </div>
        <div style="padding:8px;background:rgba(239,68,68,.08);border-radius:8px">
          <div style="color:#f87171;font-weight:600">실패 그룹 r</div>
          <div style="font-family:'JetBrains Mono';font-size:16px;font-weight:700;color:#ef4444">${{rF.toFixed(4)}}</div>
        </div>
      </div>
    </div>
    <div style="background:var(--card);border:1px solid var(--border);border-radius:12px;padding:18px">
      <div style="font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:1px;margin-bottom:10px">기초 통계</div>
      <table style="width:100%;font-size:12px">
        <tr style="border-bottom:1px solid rgba(255,255,255,.06)">
          <th style="background:transparent;padding:6px 8px;font-size:10px"></th>
          <th style="background:transparent;padding:6px 8px;font-size:10px;color:#60a5fa">성공</th>
          <th style="background:transparent;padding:6px 8px;font-size:10px;color:#f87171">실패</th>
        </tr>
        <tr style="border-bottom:1px solid rgba(255,255,255,.04)">
          <td style="padding:6px 8px;color:var(--muted)">${{VARS[xi]}} 평균</td>
          <td style="padding:6px 8px;font-family:'JetBrains Mono';color:#4f8ef7">${{avg(sXdata).toFixed(2)}}</td>
          <td style="padding:6px 8px;font-family:'JetBrains Mono';color:#ef4444">${{avg(fXdata).toFixed(2)}}</td>
        </tr>
        <tr style="border-bottom:1px solid rgba(255,255,255,.04)">
          <td style="padding:6px 8px;color:var(--muted)">${{VARS[xi]}} 표준편차</td>
          <td style="padding:6px 8px;font-family:'JetBrains Mono';color:#4f8ef7">${{std(sXdata).toFixed(2)}}</td>
          <td style="padding:6px 8px;font-family:'JetBrains Mono';color:#ef4444">${{std(fXdata).toFixed(2)}}</td>
        </tr>
        <tr style="border-bottom:1px solid rgba(255,255,255,.04)">
          <td style="padding:6px 8px;color:var(--muted)">${{VARS[yi]}} 평균</td>
          <td style="padding:6px 8px;font-family:'JetBrains Mono';color:#4f8ef7">${{avg(sYdata).toFixed(2)}}</td>
          <td style="padding:6px 8px;font-family:'JetBrains Mono';color:#ef4444">${{avg(fYdata).toFixed(2)}}</td>
        </tr>
        <tr>
          <td style="padding:6px 8px;color:var(--muted)">${{VARS[yi]}} 표준편차</td>
          <td style="padding:6px 8px;font-family:'JetBrains Mono';color:#4f8ef7">${{std(sYdata).toFixed(2)}}</td>
          <td style="padding:6px 8px;font-family:'JetBrains Mono';color:#ef4444">${{std(fYdata).toFixed(2)}}</td>
        </tr>
      </table>
      <div style="margin-top:12px;padding:10px;background:rgba(251,191,36,.06);border-radius:8px;font-size:11px;color:#a09060">
        ${{Math.abs(r) > 0.5
          ? VARS[xi] + '와(과) ' + VARS[yi] + '는 유의미한 상관관계가 있어 하나의 변수를 조절하면 다른 변수에도 영향을 줄 수 있습니다.'
          : Math.abs(r) > 0.3
          ? VARS[xi] + '와(과) ' + VARS[yi] + '는 약간의 상관관계가 있으나, 독립적으로 관리해도 무방합니다.'
          : VARS[xi] + '와(과) ' + VARS[yi] + '는 독립적인 변수로, 서로 큰 영향을 주지 않습니다.'
        }}
      </div>
    </div>`;
}}

// 페이지 로드 시 초기 계산
window.onload = function() {{
  calcPred();
  drawScatter();
}};
</script>
</body>
</html>"""

    output_path.write_text(html, encoding='utf-8')
    print(f"  [OK] HTML 보고서 저장: {output_path}")


# ─────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────
def main():
    print("\n" + "="*54)
    print("  [START] 방사형 데이터 분석기 v2 시작")
    print("="*54)

    OUTPUT_DIR.mkdir(exist_ok=True)

    # 파일 탐색
    print("\n[1/4] 엑셀 파일 자동 인식...")
    files = sorted(glob.glob("*.xlsx") + glob.glob("*.xls"))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    non_sample = [f for f in files if "sample_data" not in f]
    if non_sample:
        chosen = non_sample[0]
    elif files:
        chosen = files[0]
    else:
        sample = Path("sample_data.xlsx")
        print("  -> 파일 없음. 샘플 데이터 자동 생성...")
        create_sample_excel(sample)
        chosen = str(sample)
    print(f"  [OK] 파일: {chosen}")

    # 파싱
    print("\n[2/4] 데이터 구조 자동 분석...")
    df_vars, labels, var_names, base_rate = load_and_parse(chosen)

    # 성공 확률 계산
    print("\n[3/4] 성공 확률 계산...")
    axis_probs = calc_axis_probs(df_vars, labels)
    for v in var_names:
        print(f"  {v:20s} -> {axis_probs[v]['prob']:3d}%  "
              f"(성공 avg={axis_probs[v]['s_mean']:.1f}, "
              f"실패 avg={axis_probs[v]['f_mean']:.1f})")

    # 예측 데모
    demo_vals = [axis_probs[v]['s_mean'] for v in var_names]  # 성공 평균값으로 테스트
    demo_prob = predict_success_prob(df_vars, labels, demo_vals)
    print(f"\n  [데모] 성공 평균값 입력시 예측 성공 확률: {demo_prob}%")

    # 시각화 & 보고서 생성
    print("\n[4/4] 시각화 & 보고서 생성 (2종)...")
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    OUTPUT_DIR.mkdir(exist_ok=True)

    charts = {
        'corr':     OUTPUT_DIR / f"correlation_{ts}.png",
        'scatter':  OUTPUT_DIR / f"scatter_{ts}.png",
    }
    htm = OUTPUT_DIR / f"report_{ts}.html"

    draw_correlation(df_vars, labels, axis_probs, base_rate, charts['corr'])
    draw_scatter_matrix(df_vars, labels, axis_probs, base_rate, charts['scatter'])
    generate_html_report(df_vars, labels, axis_probs, base_rate,
                         charts, htm, chosen)

    # 상관관계 요약 출력
    print("\n  [상관관계 요약] 주요 변수 쌍:")
    corr_m = df_vars.corr()
    pairs_summary = []
    for i in range(len(var_names)):
        for j in range(i+1, len(var_names)):
            r = float(corr_m.iloc[i, j])
            if abs(r) > 0.3:
                pairs_summary.append((abs(r), r, var_names[i], var_names[j]))
    pairs_summary.sort(reverse=True)
    for _, r, vx, vy in pairs_summary[:10]:
        tag = "강" if abs(r) > 0.5 else "중"
        print(f"    [{tag}] {vx} ↔ {vy}: r={r:+.3f}")
    if not pairs_summary:
        print("    (|r| > 0.3인 유의미한 상관관계 쌍 없음)")

    print("\n" + "="*54)
    print(f"  [완료] output/ 폴더 확인 (차트 2종 + HTML 보고서)")
    print(f"  기본 성공률: {base_rate}%")
    print("="*54 + "\n")

    webbrowser.open(str(htm.resolve()))


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n[오류] {e}")
        import traceback; traceback.print_exc()
    input("\n[Enter] 키를 누르면 창이 닫힙니다...")

